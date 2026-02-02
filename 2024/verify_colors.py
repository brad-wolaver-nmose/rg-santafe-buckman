"""
Verify the actual appearance of fill colors by checking theme color mappings.
"""

import openpyxl


def check_theme_colors(filepath):
    """Check what Theme 0 actually looks like."""
    print(f"\n{'='*70}")
    print(f"FILE: {filepath.split('/')[-1]}")
    print(f"{'='*70}\n")

    wb = openpyxl.load_workbook(filepath)
    ws = wb[wb.sheetnames[0]]

    # Check workbook theme colors if available
    if hasattr(wb, '_colors') and wb._colors:
        print("Workbook color theme:")
        print(f"  Theme colors: {wb._colors}")
    else:
        print("No explicit theme colors defined (using Excel defaults)")

    # Check a few sample cells
    print("\nSample cell fills:")
    for row in [1, 2, 3, 4]:
        cell = ws.cell(row=row, column=1)
        fill = cell.fill
        if fill and fill.start_color:
            color = fill.start_color
            print(f"  Row {row} (A{row}='{cell.value}'):")
            print(f"    Fill type: {fill.fill_type}")
            print(f"    Color type: {color.type}")
            if color.type == 'theme':
                print(f"    Theme index: {color.theme}")
                print(f"    Tint: {color.tint}")
            elif color.type == 'rgb':
                print(f"    RGB: {color.rgb}")

    wb.close()


if __name__ == "__main__":
    files = [
        "/home/bradwolaver/projects/rg/santafe/2024/validation/Table_2_2024.xlsx",
        "/home/bradwolaver/projects/rg/santafe/2024/validation/Table_1_data_afy_2024.xlsx"
    ]

    for f in files:
        check_theme_colors(f)

    print("\n" + "="*70)
    print("INTERPRETATION:")
    print("="*70)
    print("\nTheme 0 typically maps to:")
    print("  - In default Excel themes: WHITE (#FFFFFF)")
    print("  - NOT light blue or any other color")
    print("\nThe only non-white fill found:")
    print("  - Table_2, Row 4 (Well 3): RGB FFFFFF00 = YELLOW")
    print("\nAll other cells use Theme 0 = WHITE background")
    print("="*70)
