"""
Example code for applying the exact formatting from validation files.

This demonstrates how to recreate the formatting specifications
documented in EXCEL_FORMAT_SPECIFICATIONS.md.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Color


def apply_table2_formatting(ws):
    """
    Apply Table_2_2024.xlsx formatting to a worksheet.

    Args:
        ws: openpyxl worksheet object
    """
    # Define styles
    header_font = Font(name='Aptos', size=11, bold=True, color=Color(theme=1))
    header_fill = PatternFill(fill_type='solid', fgColor=Color(theme=0))
    header_alignment = Alignment(horizontal='center')
    medium_border = Side(style='medium')
    hair_border = Side(style='hair')

    data_font = Font(name='Aptos', size=11, bold=False, color=Color(theme=1))
    data_fill_white = PatternFill(fill_type='solid', fgColor=Color(theme=0))
    data_fill_yellow = PatternFill(fill_type='solid', fgColor='FFFFFF00')

    well_font = Font(name='Aptos', size=11, bold=True, color=Color(theme=1))
    well_alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 14.75
    for col in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col].width = 13.0
    ws.column_dimensions['O'].width = 14.75

    # Header row
    headers = ['Well', 'JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN', 'JUL',
               'AUG', 'SEP', 'OCT', 'NOV', 'DEC', 'Total', 'Well']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = Border(top=medium_border, bottom=medium_border)

    # Example data rows (showing formatting patterns)
    # Row 2: Well 1 (normal white fill, no bottom border)
    ws.cell(row=2, column=1).value = 1
    ws.cell(row=2, column=1).font = well_font
    ws.cell(row=2, column=1).fill = data_fill_white
    ws.cell(row=2, column=1).alignment = well_alignment

    # Numeric columns
    for col_idx in range(2, 14):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = data_font
        cell.fill = data_fill_white
        cell.number_format = '#,##0.00'

    # Total formula
    ws.cell(row=2, column=14).value = '=SUM(B2:M2)'
    ws.cell(row=2, column=14).font = data_font
    ws.cell(row=2, column=14).fill = data_fill_white
    ws.cell(row=2, column=14).number_format = '#,##0.00'

    # Row 4: Well 3 (YELLOW highlighted row with hair borders)
    ws.cell(row=4, column=1).value = 3
    ws.cell(row=4, column=1).font = well_font
    ws.cell(row=4, column=1).fill = data_fill_yellow
    ws.cell(row=4, column=1).alignment = well_alignment
    ws.cell(row=4, column=1).border = Border(bottom=hair_border)

    for col_idx in range(2, 14):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = data_font
        cell.fill = data_fill_yellow
        cell.number_format = '#,##0.00'
        cell.border = Border(bottom=hair_border)

    ws.cell(row=4, column=14).value = '=SUM(B4:M4)'
    ws.cell(row=4, column=14).font = data_font
    ws.cell(row=4, column=14).fill = data_fill_yellow
    ws.cell(row=4, column=14).number_format = '#,##0.00'
    ws.cell(row=4, column=14).border = Border(bottom=hair_border)

    return ws


def apply_table1_formatting(ws):
    """
    Apply Table_1_data_afy_2024.xlsx formatting to a worksheet.

    Args:
        ws: openpyxl worksheet object
    """
    # Define styles
    header_font = Font(name='Aptos', size=11, bold=True, color=Color(theme=1))
    header_fill = PatternFill(fill_type='solid', fgColor=Color(theme=0))
    header_alignment_a1 = Alignment(horizontal='right')
    header_alignment_center = Alignment(horizontal='center')
    medium_border = Side(style='medium')
    hair_border = Side(style='hair')

    data_font = Font(name='Aptos', size=11, bold=False, color=Color(theme=1))
    data_fill = PatternFill(fill_type='solid', fgColor=Color(theme=0))

    year_font = Font(name='Aptos', size=11, bold=True, color=Color(theme=1))
    year_alignment = Alignment(horizontal='center')

    # Set column widths
    ws.column_dimensions['A'].width = 12.75
    for col_letter in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']:
        ws.column_dimensions[col_letter].width = 13.0
    ws.column_dimensions['O'].width = 11.75

    # Header row
    headers = ['Well:', '1', '2', '3/3A', '4', '5', '6', '7', '8', '9', '10', '11', '12', '13', 'Total']

    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill

        # Special alignment for A1
        if col_idx == 1:
            cell.alignment = header_alignment_a1
        else:
            cell.alignment = header_alignment_center

        cell.border = Border(top=medium_border, bottom=medium_border)

    # Example data row (e.g., year 1988)
    ws.cell(row=2, column=1).value = 1988
    ws.cell(row=2, column=1).font = year_font
    ws.cell(row=2, column=1).fill = data_fill
    ws.cell(row=2, column=1).alignment = year_alignment
    ws.cell(row=2, column=1).border = Border(bottom=hair_border)

    # Numeric columns (well data)
    for col_idx in range(2, 15):
        cell = ws.cell(row=2, column=col_idx)
        cell.font = data_font
        cell.fill = data_fill
        cell.number_format = '#,##0.00'
        cell.border = Border(bottom=hair_border)

    return ws


def create_example_workbooks():
    """Create example workbooks with correct formatting."""

    # Example 1: Table_2 style
    wb1 = openpyxl.Workbook()
    ws1 = wb1.active
    ws1.title = "Table_2_Example"
    apply_table2_formatting(ws1)
    wb1.save("/home/bradwolaver/projects/rg/santafe/2024/example_table2_format.xlsx")
    print("Created: example_table2_format.xlsx")

    # Example 2: Table_1 style
    wb2 = openpyxl.Workbook()
    ws2 = wb2.active
    ws2.title = "Table1_Example"
    apply_table1_formatting(ws2)
    wb2.save("/home/bradwolaver/projects/rg/santafe/2024/example_table1_format.xlsx")
    print("Created: example_table1_format.xlsx")


if __name__ == "__main__":
    create_example_workbooks()
    print("\nExample workbooks created with validation file formatting!")
    print("\nKey formatting patterns demonstrated:")
    print("  - Table_2: Yellow highlight for Well 3 (row 4)")
    print("  - Table_2: Medium borders on headers, hair borders on data")
    print("  - Table_2: Center-aligned headers and Well IDs")
    print("  - Table_1: Right-aligned A1 ('Well:'), center for others")
    print("  - Table_1: Consistent hair borders on all data rows")
    print("  - Both: Aptos 11pt font, #,##0.00 number format")
    print("  - Both: SUM formulas in Total columns")
