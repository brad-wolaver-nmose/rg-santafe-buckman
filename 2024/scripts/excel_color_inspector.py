"""
Inspect actual color values in Excel files.
"""

import openpyxl


def inspect_colors(filepath):
    """Inspect colors in Excel file."""
    print(f"\n{'='*80}")
    print(f"FILE: {filepath.split('/')[-1]}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=False)
    ws = wb[wb.sheetnames[0]]  # First sheet

    print("Header Row (Row 1) Colors:")
    cell = ws.cell(row=1, column=1)
    print(f"\nA1 ('{cell.value}'):")
    print(f"  Font color object: {cell.font.color}")
    print(f"  Font color type: {type(cell.font.color)}")
    if cell.font.color:
        print(f"  Font color.__dict__: {cell.font.color.__dict__}")

    print(f"\n  Fill object: {cell.fill}")
    print(f"  Fill type: {cell.fill.fill_type}")
    if cell.fill.start_color:
        print(f"  Fill start_color object: {cell.fill.start_color}")
        print(f"  Fill start_color type: {type(cell.fill.start_color)}")
        print(f"  Fill start_color.__dict__: {cell.fill.start_color.__dict__}")

    print(f"\n\nData Rows (Rows 2-6) Colors:")
    for row_num in range(2, 7):
        cell = ws.cell(row=row_num, column=1)
        print(f"\nA{row_num} (Well '{cell.value}'):")

        if cell.fill and cell.fill.fill_type:
            print(f"  Fill type: {cell.fill.fill_type}")
            if cell.fill.start_color:
                print(f"  Fill start_color.__dict__: {cell.fill.start_color.__dict__}")

    wb.close()


if __name__ == "__main__":
    files = [
        "/home/bradwolaver/projects/rg/santafe/2024/validation/Table_2_2024.xlsx",
        "/home/bradwolaver/projects/rg/santafe/2024/validation/Table_1_data_afy_2024.xlsx"
    ]

    for filepath in files:
        try:
            inspect_colors(filepath)
        except Exception as e:
            print(f"Error: {e}")
            import traceback
            traceback.print_exc()
