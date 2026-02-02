"""
Concise Excel formatting summary for validation files.

Extracts key formatting details in a compact, structured format.
"""

import openpyxl
from openpyxl.utils import get_column_letter


def get_color_str(color_obj):
    """Extract color string from openpyxl color object."""
    if not color_obj:
        return None
    if hasattr(color_obj, 'rgb') and color_obj.rgb:
        return color_obj.rgb
    if hasattr(color_obj, 'index'):
        return color_obj.index
    return None


def summarize_excel(filepath):
    """Create compact summary of Excel formatting."""
    print(f"\n{'='*80}")
    print(f"FILE: {filepath.split('/')[-1]}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSHEET: {sheet_name}")
        print("-" * 60)

        # Dimensions
        print(f"\nDIMENSIONS: {ws.max_row} rows x {ws.max_column} columns")

        # Frozen panes
        if ws.freeze_panes:
            print(f"FROZEN PANES: {ws.freeze_panes}")

        # Merged cells
        if ws.merged_cells:
            print(f"\nMERGED CELLS: {len(ws.merged_cells.ranges)} ranges")
            for merged in list(ws.merged_cells.ranges)[:5]:
                print(f"  - {merged}")
            if len(ws.merged_cells.ranges) > 5:
                print(f"  ... and {len(ws.merged_cells.ranges) - 5} more")

        # Column widths
        print(f"\nCOLUMN WIDTHS:")
        widths = {}
        for col_num in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_num)
            width = ws.column_dimensions[col_letter].width
            if width not in widths:
                widths[width] = []
            widths[width].append(col_letter)
        for width, cols in sorted(widths.items()):
            if len(cols) <= 3:
                print(f"  {width}: {', '.join(cols)}")
            else:
                print(f"  {width}: {cols[0]}-{cols[-1]} ({len(cols)} columns)")

        # Header row analysis
        print(f"\nHEADER ROW (Row 1):")
        header_row = []
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            if cell.value:
                header_row.append(str(cell.value))
        print(f"  Columns: {', '.join(header_row)}")

        # Get unique formatting patterns from first cell
        cell = ws.cell(row=1, column=1)
        print(f"\n  Header formatting (A1):")
        print(f"    Font: {cell.font.name}, Size: {cell.font.size}, Bold: {cell.font.bold}")

        font_color = get_color_str(cell.font.color)
        if font_color:
            print(f"    Font Color: {font_color}")

        if cell.fill and cell.fill.start_color:
            fill_color = get_color_str(cell.fill.start_color)
            if fill_color and fill_color != '00000000':
                print(f"    Fill: {cell.fill.fill_type}, Color: {fill_color}")

        if cell.alignment:
            print(f"    Alignment: H={cell.alignment.horizontal}, V={cell.alignment.vertical}")

        if cell.border:
            borders = []
            if cell.border.top and cell.border.top.style:
                borders.append(f"Top: {cell.border.top.style}")
            if cell.border.bottom and cell.border.bottom.style:
                borders.append(f"Bottom: {cell.border.bottom.style}")
            if borders:
                print(f"    Borders: {', '.join(borders)}")

        # Data row analysis (row 2)
        print(f"\nDATA ROW FORMAT (Row 2):")
        cell = ws.cell(row=2, column=1)
        print(f"  First column (A2): Value='{cell.value}'")
        print(f"    Font: Bold={cell.font.bold}, Size={cell.font.size}")
        if cell.alignment:
            print(f"    Alignment: H={cell.alignment.horizontal}")

        # Check numeric columns
        for col_num in range(2, min(ws.max_column + 1, 5)):
            cell = ws.cell(row=2, column=col_num)
            col_letter = get_column_letter(col_num)
            if cell.value is not None:
                print(f"  {col_letter}2: Value={cell.value}, Type={type(cell.value).__name__}")
                print(f"    Number Format: '{cell.number_format}'")
                break

        # Check for special formatting (colored rows, etc.)
        print(f"\nSPECIAL FORMATTING:")
        special_rows = []
        for row_num in range(2, min(ws.max_row + 1, 20)):
            cell = ws.cell(row=row_num, column=1)
            if cell.fill and cell.fill.start_color:
                fill_color = get_color_str(cell.fill.start_color)
                if fill_color and fill_color not in ['00000000', None]:
                    special_rows.append((row_num, fill_color))

        if special_rows:
            print(f"  Colored rows detected:")
            for row_num, color in special_rows[:5]:
                cell_val = ws.cell(row=row_num, column=1).value
                print(f"    Row {row_num} (Well {cell_val}): Fill color {color}")
        else:
            print(f"  No special colored rows detected in first 20 rows")

        # Border patterns
        print(f"\nBORDER PATTERNS:")
        border_styles = set()
        for row_num in range(1, min(ws.max_row + 1, 10)):
            cell = ws.cell(row=row_num, column=1)
            if cell.border and cell.border.bottom and cell.border.bottom.style:
                border_styles.add((row_num, cell.border.bottom.style))
        if border_styles:
            for row_num, style in sorted(border_styles):
                print(f"  Row {row_num}: Bottom border = {style}")

        # Summary row (usually last row)
        if ws.max_row > 2:
            print(f"\nLAST ROW (Row {ws.max_row}):")
            cell = ws.cell(row=ws.max_row, column=1)
            print(f"  A{ws.max_row}: '{cell.value}'")
            if cell.font.bold:
                print(f"  Font: BOLD")

            # Check for sum formulas
            has_formula = False
            for col_num in range(2, min(ws.max_column + 1, 5)):
                cell = ws.cell(row=ws.max_row, column=col_num)
                if isinstance(cell.value, str) and cell.value.startswith('='):
                    print(f"  {get_column_letter(col_num)}{ws.max_row}: Formula = {cell.value}")
                    has_formula = True
                    break

    wb.close()


if __name__ == "__main__":
    files = [
        "/home/bradwolaver/projects/rg/santafe/2024/validation/Table_2_2024.xlsx",
        "/home/bradwolaver/projects/rg/santafe/2024/validation/Table_1_data_afy_2024.xlsx"
    ]

    for filepath in files:
        try:
            summarize_excel(filepath)
        except Exception as e:
            print(f"\nError analyzing {filepath}: {e}")
            import traceback
            traceback.print_exc()

    print(f"\n{'='*80}\n")
