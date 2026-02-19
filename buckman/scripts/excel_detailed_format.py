"""
Detailed Excel formatting analysis with proper color handling.
"""

import openpyxl
from openpyxl.utils import get_column_letter


def get_color_info(color_obj):
    """Extract color information from openpyxl color object."""
    if not color_obj:
        return None
    try:
        # Try RGB first
        if hasattr(color_obj, 'rgb') and color_obj.rgb:
            return f"RGB:{color_obj.rgb}"
        # Then theme
        if hasattr(color_obj, 'theme') and color_obj.theme is not None:
            return f"Theme:{color_obj.theme}"
        # Then indexed
        if hasattr(color_obj, 'indexed') and color_obj.indexed is not None:
            return f"Indexed:{color_obj.indexed}"
        # Fallback to index
        if hasattr(color_obj, 'index'):
            return f"Index:{color_obj.index}"
    except Exception:
        pass
    return "Unknown"


def analyze_file(filepath):
    """Detailed analysis of Excel file formatting."""
    filename = filepath.split('/')[-1]
    print(f"\n{'#'*80}")
    print(f"# {filename}")
    print(f"{'#'*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=False)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n## SHEET: {sheet_name}")
        print("="*70)

        # Basic structure
        print("\n### Structure")
        print(f"- Dimensions: {ws.max_row} rows × {ws.max_column} columns")
        print(f"- Range: A1:{get_column_letter(ws.max_column)}{ws.max_row}")

        # Frozen panes
        if ws.freeze_panes:
            print(f"- Frozen panes: {ws.freeze_panes}")

        # Merged cells
        if ws.merged_cells:
            print(f"\n### Merged Cells: {len(ws.merged_cells.ranges)}")
            for merged in ws.merged_cells.ranges:
                print(f"  - {merged}")

        # Column widths
        print("\n### Column Widths")
        width_groups = {}
        for col_num in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_num)
            width = ws.column_dimensions[col_letter].width
            if width not in width_groups:
                width_groups[width] = []
            width_groups[width].append((col_num, col_letter))

        for width in sorted(width_groups.keys()):
            cols = width_groups[width]
            if len(cols) == 1:
                print(f"  - {cols[0][1]}: {width}")
            elif len(cols) <= 5:
                col_str = ', '.join([c[1] for c in cols])
                print(f"  - {col_str}: {width}")
            else:
                print(f"  - {cols[0][1]}-{cols[-1][1]} ({len(cols)} cols): {width}")

        # Header row
        print("\n### Header Row (Row 1)")
        print("Columns:")
        for col_num in range(1, ws.max_column + 1):
            cell = ws.cell(row=1, column=col_num)
            col_letter = get_column_letter(col_num)
            if cell.value:
                print(f"  {col_letter}: '{cell.value}'")

        # Header formatting (from A1)
        cell = ws.cell(row=1, column=1)
        print("\nHeader Format (A1 example):")
        print("  Font:")
        print(f"    - Name: {cell.font.name}")
        print(f"    - Size: {cell.font.size}")
        print(f"    - Bold: {cell.font.bold}")
        print(f"    - Italic: {cell.font.italic}")
        if cell.font.color:
            print(f"    - Color: {get_color_info(cell.font.color)}")

        if cell.fill and cell.fill.fill_type:
            print("  Fill:")
            print(f"    - Type: {cell.fill.fill_type}")
            if cell.fill.start_color:
                print(f"    - Start Color: {get_color_info(cell.fill.start_color)}")
            if cell.fill.end_color and get_color_info(cell.fill.end_color) != "Index:00000000":
                print(f"    - End Color: {get_color_info(cell.fill.end_color)}")

        if cell.alignment:
            print("  Alignment:")
            print(f"    - Horizontal: {cell.alignment.horizontal}")
            print(f"    - Vertical: {cell.alignment.vertical}")
            if cell.alignment.wrap_text:
                print("    - Wrap Text: True")

        if cell.border:
            borders = []
            if cell.border.left and cell.border.left.style:
                borders.append(f"Left: {cell.border.left.style}")
            if cell.border.right and cell.border.right.style:
                borders.append(f"Right: {cell.border.right.style}")
            if cell.border.top and cell.border.top.style:
                borders.append(f"Top: {cell.border.top.style}")
            if cell.border.bottom and cell.border.bottom.style:
                borders.append(f"Bottom: {cell.border.bottom.style}")
            if borders:
                print(f"  Borders: {', '.join(borders)}")

        # Data rows - sample first few
        print("\n### Data Rows (Rows 2-6)")
        for row_num in range(2, min(7, ws.max_row + 1)):
            print(f"\nRow {row_num}:")

            # First column (usually well/year ID)
            cell = ws.cell(row=row_num, column=1)
            print(f"  A{row_num}: '{cell.value}'")
            print(f"    - Number Format: '{cell.number_format}'")
            print(f"    - Bold: {cell.font.bold}")
            if cell.alignment and cell.alignment.horizontal:
                print(f"    - Align: {cell.alignment.horizontal}")

            # Check for fill color
            if cell.fill and cell.fill.start_color:
                fill_info = get_color_info(cell.fill.start_color)
                if fill_info and "00000000" not in fill_info:
                    print(f"    - Fill: {cell.fill.fill_type}, Color: {fill_info}")

            # Border
            if cell.border and cell.border.bottom and cell.border.bottom.style:
                print(f"    - Bottom Border: {cell.border.bottom.style}")

            # Second column (first numeric)
            cell = ws.cell(row=row_num, column=2)
            if cell.value is not None:
                print(f"  B{row_num}: {cell.value} ({type(cell.value).__name__})")
                print(f"    - Number Format: '{cell.number_format}'")

            # Check for formula in Total column
            total_col = 14 if sheet_name == "Table_2_2024" else 15
            if total_col <= ws.max_column:
                cell = ws.cell(row=row_num, column=total_col)
                col_letter = get_column_letter(total_col)
                if isinstance(cell.value, str) and '=' in cell.value:
                    print(f"  {col_letter}{row_num}: {cell.value}")

        # Last row
        if ws.max_row > 10:
            print(f"\n### Last Row (Row {ws.max_row})")
            cell = ws.cell(row=ws.max_row, column=1)
            print(f"  A{ws.max_row}: '{cell.value}'")
            if cell.font.bold:
                print("    - Font: BOLD")

            # Check for formulas
            for col_num in range(2, min(ws.max_column + 1, 6)):
                cell = ws.cell(row=ws.max_row, column=col_num)
                if cell.value is not None:
                    col_letter = get_column_letter(col_num)
                    if isinstance(cell.value, str) and '=' in cell.value:
                        print(f"  {col_letter}{ws.max_row}: {cell.value}")
                    else:
                        print(f"  {col_letter}{ws.max_row}: {cell.value}")

        # Number formats in use
        print("\n### Number Formats Used")
        formats_used = set()
        for row_num in range(1, min(ws.max_row + 1, 20)):
            for col_num in range(1, ws.max_column + 1):
                cell = ws.cell(row=row_num, column=col_num)
                if cell.value is not None and cell.number_format != 'General':
                    formats_used.add(cell.number_format)

        for fmt in sorted(formats_used):
            print(f"  - '{fmt}'")

        print("\n" + "="*70)

    wb.close()


if __name__ == "__main__":
    files = [
        "/home/bradwolaver/projects/rg/santafe/buckman/validation/Table_2_2024.xlsx",
        "/home/bradwolaver/projects/rg/santafe/buckman/validation/Table_1_data_afy_2024.xlsx"
    ]

    for filepath in files:
        try:
            analyze_file(filepath)
        except Exception as e:
            print(f"\nError analyzing {filepath}: {e}")
            import traceback
            traceback.print_exc()
