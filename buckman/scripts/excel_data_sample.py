"""
Extract sample data from validation Excel files to show structure.
"""

import openpyxl
from openpyxl.utils import get_column_letter


def show_data_sample(filepath):
    """Show first few rows of data."""
    filename = filepath.split('/')[-1]
    print(f"\n{'#'*80}")
    print(f"# {filename}")
    print(f"{'#'*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=True)  # data_only=True to see formula results
    ws = wb[wb.sheetnames[0]]

    # Print header and first 5 data rows
    for row_num in range(1, min(8, ws.max_row + 1)):
        row_data = []
        for col_num in range(1, min(ws.max_column + 1, 15)):  # First 15 columns
            cell = ws.cell(row=row_num, column=col_num)
            value = cell.value if cell.value is not None else ""
            row_data.append(str(value))

        if row_num == 1:
            print("HEADER ROW:")
            print("  " + " | ".join(row_data))
            print("  " + "-" * 60)
        else:
            print(f"Row {row_num:2d}: " + " | ".join(row_data))

    wb.close()


if __name__ == "__main__":
    files = [
        "/home/bradwolaver/projects/rg/santafe/buckman/validation/Table_2_2024.xlsx",
        "/home/bradwolaver/projects/rg/santafe/buckman/validation/Table_1_data_afy_2024.xlsx"
    ]

    for filepath in files:
        show_data_sample(filepath)
