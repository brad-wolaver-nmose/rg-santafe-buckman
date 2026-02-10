"""
Comprehensive Excel formatting analyzer for validation files.

This script examines all formatting details from Excel files including:
- Sheet structure (names, dimensions)
- Column headers and data
- Cell formatting (number formats, fonts, fills, borders, alignment)
- Merged cells
- Column widths and row heights
- Frozen panes
- Conditional formatting
"""

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Fill, Border, Alignment
import sys


def analyze_excel_file(filepath):
    """
    Analyze all formatting details of an Excel file.

    Args:
        filepath: Path to Excel file

    Returns:
        dict: Comprehensive formatting analysis
    """
    print(f"\n{'='*80}")
    print(f"ANALYZING: {filepath}")
    print(f"{'='*80}\n")

    wb = openpyxl.load_workbook(filepath, data_only=False)

    analysis = {
        'filename': filepath,
        'sheet_names': wb.sheetnames,
        'sheets': {}
    }

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\n{'*'*60}")
        print(f"SHEET: {sheet_name}")
        print(f"{'*'*60}\n")

        # Get dimensions
        min_row = ws.min_row
        max_row = ws.max_row
        min_col = ws.min_column
        max_col = ws.max_column

        print(f"Dimensions: {max_row} rows x {max_col} columns")
        print(f"Range: {get_column_letter(min_col)}{min_row}:{get_column_letter(max_col)}{max_row}\n")

        # Frozen panes
        if ws.freeze_panes:
            print(f"Frozen panes: {ws.freeze_panes}\n")

        # Merged cells
        if ws.merged_cells:
            print(f"Merged cells ({len(ws.merged_cells.ranges)}):")
            for merged_range in ws.merged_cells.ranges:
                print(f"  {merged_range}")
            print()

        # Column widths
        print("Column widths:")
        for col_num in range(min_col, max_col + 1):
            col_letter = get_column_letter(col_num)
            width = ws.column_dimensions[col_letter].width
            print(f"  {col_letter}: {width}")
        print()

        # Row heights (first 20 rows)
        print("Row heights (first 20):")
        for row_num in range(min_row, min(max_row + 1, min_row + 20)):
            height = ws.row_dimensions[row_num].height
            if height:
                print(f"  Row {row_num}: {height}")
        print()

        # Analyze cells - headers and sample data
        print("="*60)
        print("CELL ANALYSIS")
        print("="*60)

        # Check first 10 rows for structure
        for row_num in range(min_row, min(max_row + 1, min_row + 10)):
            print(f"\n--- Row {row_num} ---")
            for col_num in range(min_col, max_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                col_letter = get_column_letter(col_num)

                # Only print if cell has value or notable formatting
                if cell.value is not None or cell.font.bold or cell.fill.start_color.index != '00000000':
                    print(f"\nCell {col_letter}{row_num}:")
                    print(f"  Value: {cell.value}")
                    print(f"  Type: {type(cell.value).__name__}")

                    # Number format
                    if cell.number_format:
                        print(f"  Number Format: {cell.number_format}")

                    # Font
                    if cell.font:
                        font_info = []
                        if cell.font.bold:
                            font_info.append("BOLD")
                        if cell.font.italic:
                            font_info.append("ITALIC")
                        if cell.font.name:
                            font_info.append(f"Font: {cell.font.name}")
                        if cell.font.size:
                            font_info.append(f"Size: {cell.font.size}")
                        if cell.font.color and cell.font.color.rgb:
                            font_info.append(f"Color: {cell.font.color.rgb}")
                        if font_info:
                            print(f"  Font: {', '.join(font_info)}")

                    # Fill
                    if cell.fill and cell.fill.start_color.index != '00000000':
                        fill_type = cell.fill.fill_type
                        start_color = cell.fill.start_color.rgb if cell.fill.start_color.rgb else cell.fill.start_color.index
                        print(f"  Fill: {fill_type}, Color: {start_color}")
                        if cell.fill.end_color and cell.fill.end_color.index != '00000000':
                            end_color = cell.fill.end_color.rgb if cell.fill.end_color.rgb else cell.fill.end_color.index
                            print(f"    End Color: {end_color}")

                    # Alignment
                    if cell.alignment:
                        align_info = []
                        if cell.alignment.horizontal:
                            align_info.append(f"H:{cell.alignment.horizontal}")
                        if cell.alignment.vertical:
                            align_info.append(f"V:{cell.alignment.vertical}")
                        if cell.alignment.wrap_text:
                            align_info.append("WRAP")
                        if cell.alignment.indent:
                            align_info.append(f"Indent:{cell.alignment.indent}")
                        if align_info:
                            print(f"  Alignment: {', '.join(align_info)}")

                    # Borders
                    if cell.border:
                        borders = []
                        if cell.border.left and cell.border.left.style:
                            borders.append(f"L:{cell.border.left.style}")
                        if cell.border.right and cell.border.right.style:
                            borders.append(f"R:{cell.border.right.style}")
                        if cell.border.top and cell.border.top.style:
                            borders.append(f"T:{cell.border.top.style}")
                        if cell.border.bottom and cell.border.bottom.style:
                            borders.append(f"B:{cell.border.bottom.style}")
                        if borders:
                            print(f"  Borders: {', '.join(borders)}")

        # Sample a few middle rows
        if max_row > 20:
            print(f"\n\n--- Sample from middle rows (row {max_row//2}) ---")
            row_num = max_row // 2
            for col_num in range(min_col, max_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                col_letter = get_column_letter(col_num)
                if cell.value is not None:
                    print(f"\n{col_letter}{row_num}: {cell.value} (Format: {cell.number_format})")

        # Check last few rows
        print(f"\n\n--- Last 3 rows ---")
        for row_num in range(max(min_row, max_row - 2), max_row + 1):
            print(f"\nRow {row_num}:")
            for col_num in range(min_col, max_col + 1):
                cell = ws.cell(row=row_num, column=col_num)
                col_letter = get_column_letter(col_num)
                if cell.value is not None:
                    print(f"  {col_letter}: {cell.value} (Format: {cell.number_format})")

        # Conditional formatting
        if ws.conditional_formatting:
            print(f"\n\nConditional Formatting Rules: {len(ws.conditional_formatting._cf_rules)}")
            for rule_range, rules in ws.conditional_formatting._cf_rules.items():
                print(f"  Range: {rule_range}")
                for rule in rules:
                    print(f"    Type: {rule.type}")
                    print(f"    Rule: {rule}")

    wb.close()
    return analysis


if __name__ == "__main__":
    files = [
        "/home/bradwolaver/projects/rg/santafe/buckman/validation/Table_2_2024.xlsx",
        "/home/bradwolaver/projects/rg/santafe/buckman/validation/Table_1_data_afy_2024.xlsx"
    ]

    for filepath in files:
        try:
            analyze_excel_file(filepath)
        except Exception as e:
            print(f"Error analyzing {filepath}: {e}")
            import traceback
            traceback.print_exc()
