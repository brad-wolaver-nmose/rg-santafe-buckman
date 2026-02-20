#!/usr/bin/env python3
"""
Cross-model verification: compare CY{N-1} vs CY{N} superposition values.

Parses both post-processor output files (e.g., CY2024 and CY2025) and computes
R POJOAQUE and R TESUQUE superposition acre-feet for every year 1988-2030 directly
from MODFLOW output — no historical baseline override. This reveals whether two
consecutive MODFLOW runs agree on overlapping years.

Usage:
    python3 verify_depletion.py --year 2025

Output:
    output/depletion/Table_3_verify_depletion_{YEAR-1}_{YEAR}.xlsx
"""

import argparse
import calendar
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

import stream_depletions as sd
from step4_generate_depletion_tables import (
    get_depletions_dir,
    get_output_file_prefix,
    parse_post_processor_output,
)

# Year range for comparison (Table 3 covers 1988-2030)
YEAR_START = 1988
YEAR_END = 2030

# Streams to compare
STREAMS = ["R POJOAQUE", "R TESUQUE"]
STREAM_LABELS = {
    "R POJOAQUE": "Rio Pojoaque-Nambe",
    "R TESUQUE": "Rio Tesuque",
}

# Threshold for flagging differences (acre-feet)
DIFF_THRESHOLD_AF = 0.001

# Month names matching post-processor output
MONTHS = ["jan", "feb", "mar", "apr", "may", "jun",
          "jul", "aug", "sep", "oct", "nov", "dec"]


def compute_superposition_series(
    parsed_data: dict[int, dict[str, dict[str, float]]],
    stream: str,
    year_start: int = YEAR_START,
    year_end: int = YEAR_END,
) -> dict[int, float]:
    """
    Compute annual superposition AF for a stream across all years.

    For each year, extracts 12 monthly cfs values from the parsed post-processor
    output and converts to annual acre-feet using actual calendar days (leap-year
    aware).

    Args:
        parsed_data: Output from parse_post_processor_output().
        stream: Stream identifier (e.g., "R POJOAQUE").
        year_start: First year in series.
        year_end: Last year in series (inclusive).

    Returns:
        Dict mapping year to superposition AF value.
    """
    result: dict[int, float] = {}
    for yr in range(year_start, year_end + 1):
        if yr not in parsed_data:
            print(f"  WARNING: Year {yr} not in parsed data for {stream}")
            continue
        if stream not in parsed_data[yr]:
            print(f"  WARNING: {stream} not in year {yr} data")
            continue

        # Verify all 12 months are present before extracting
        if not all(m in parsed_data[yr][stream] for m in MONTHS):
            missing = [m for m in MONTHS if m not in parsed_data[yr][stream]]
            print(f"  WARNING: Year {yr} {stream} missing months: {missing}")
            continue

        cfs_values = [parsed_data[yr][stream][m] for m in MONTHS]
        is_leap = calendar.isleap(yr)
        af = sd.cfs_monthly_to_af_annual(cfs_values, year=yr, use_leap_year=is_leap)
        result[yr] = af

    return result


def write_verification_xlsx(
    output_path: Path,
    prev_year: int,
    curr_year: int,
    data: dict[str, dict[str, dict[int, float]]],
) -> None:
    """
    Write the verification comparison xlsx.

    Args:
        output_path: Path for output xlsx file.
        prev_year: Previous model run year (e.g., 2024).
        curr_year: Current model run year (e.g., 2025).
        data: Nested dict: {stream: {"prev": {yr: af}, "curr": {yr: af}}}
    """
    wb = Workbook()
    ws = wb.active
    ws.title = f"CY{prev_year} vs CY{curr_year}"

    # Styles
    font_header = Font(name="Calibri", size=10, bold=True)
    font_normal = Font(name="Calibri", size=10)
    font_bold = Font(name="Calibri", size=10, bold=True)
    align_center = Alignment(horizontal="center", vertical="center")
    align_center_wrap = Alignment(horizontal="center", vertical="center", wrap_text=True)
    num_fmt_3 = "0.000"
    num_fmt_6 = "0.000000"
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    thin_side = Side(style="thin")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)

    # Headers (row 1)
    headers = [
        "Year",
        f"Pojoaque\nCY{prev_year}\n(AF)",
        f"Pojoaque\nCY{curr_year}\n(AF)",
        "Pojoaque\nDiff",
        f"Tesuque\nCY{prev_year}\n(AF)",
        f"Tesuque\nCY{curr_year}\n(AF)",
        "Tesuque\nDiff",
    ]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = font_header
        cell.alignment = align_center_wrap
        cell.border = thin_border

    ws.row_dimensions[1].height = 50

    # Data rows
    years = list(range(YEAR_START, YEAR_END + 1))

    # Track max diffs for summary
    max_diffs: dict[str, tuple[float, int]] = {}  # stream -> (max_abs_diff, year)

    for row_idx, yr in enumerate(years, start=2):
        # Column A: Year
        cell = ws.cell(row=row_idx, column=1, value=yr)
        cell.font = font_normal
        cell.alignment = align_center
        cell.border = thin_border

        col_offset = 1
        for stream in STREAMS:
            prev_af = data[stream]["prev"].get(yr)
            curr_af = data[stream]["curr"].get(yr)

            # Previous model value
            cell_prev = ws.cell(row=row_idx, column=col_offset + 1, value=prev_af)
            cell_prev.font = font_normal
            cell_prev.alignment = align_center
            cell_prev.number_format = num_fmt_3
            cell_prev.border = thin_border

            # Current model value
            cell_curr = ws.cell(row=row_idx, column=col_offset + 2, value=curr_af)
            cell_curr.font = font_normal
            cell_curr.alignment = align_center
            cell_curr.number_format = num_fmt_3
            cell_curr.border = thin_border

            # Diff
            if prev_af is not None and curr_af is not None:
                diff = curr_af - prev_af
                cell_diff = ws.cell(row=row_idx, column=col_offset + 3, value=diff)
                cell_diff.number_format = num_fmt_6

                # Track max diff
                abs_diff = abs(diff)
                if stream not in max_diffs or abs_diff > max_diffs[stream][0]:
                    max_diffs[stream] = (abs_diff, yr)

                # Highlight if above threshold
                if abs_diff > DIFF_THRESHOLD_AF:
                    cell_diff.fill = yellow_fill
            else:
                cell_diff = ws.cell(row=row_idx, column=col_offset + 3, value=None)

            cell_diff.font = font_normal
            cell_diff.alignment = align_center
            cell_diff.border = thin_border

            col_offset += 3

    # Summary row
    summary_row = len(years) + 2
    cell = ws.cell(row=summary_row, column=1, value="Max |Diff|")
    cell.font = font_bold
    cell.alignment = align_center
    cell.border = thin_border

    col_offset = 1
    for stream in STREAMS:
        # Leave prev/curr columns blank in summary
        for c in (col_offset + 1, col_offset + 2):
            cell = ws.cell(row=summary_row, column=c)
            cell.border = thin_border

        if stream in max_diffs:
            max_val, max_yr = max_diffs[stream]
            cell = ws.cell(row=summary_row, column=col_offset + 3, value=max_val)
            cell.font = font_bold
            cell.alignment = align_center
            cell.number_format = num_fmt_6
            cell.border = thin_border
            if max_val > DIFF_THRESHOLD_AF:
                cell.fill = yellow_fill
        col_offset += 3

    # Column widths
    widths = [8, 16, 16, 14, 16, 16, 14]
    from openpyxl.utils import get_column_letter
    for col, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(col)].width = width

    wb.save(output_path)


def main(year: int | None = None) -> int:
    """
    Compare superposition values between consecutive model runs.

    Args:
        year: Current processing year (e.g., 2025). Previous year = year - 1.

    Returns:
        0 if all historical years match within threshold, 1 if flags found.
    """
    if year is None:
        year = 2025

    prev_year = year - 1
    print(f"=== Depletion Verification: CY{prev_year} vs CY{year} ===\n")

    # Build paths to CY files
    prev_file = str(Path(get_depletions_dir(prev_year)) / get_output_file_prefix(prev_year))
    curr_file = str(Path(get_depletions_dir(year)) / get_output_file_prefix(year))

    print(f"Previous model: {prev_file}")
    print(f"Current model:  {curr_file}")

    # Check files exist
    if not Path(prev_file).exists():
        print(f"ERROR: {prev_file} not found")
        print(f"  Both CY{prev_year} and CY{year} post-processor output required.")
        print(f"  Run: python3 step4_generate_depletion_tables.py --year {prev_year}")
        return 1
    if not Path(curr_file).exists():
        print(f"ERROR: {curr_file} not found")
        print(f"  Both CY{prev_year} and CY{year} post-processor output required.")
        print(f"  Run: python3 step4_generate_depletion_tables.py --year {year}")
        return 1

    # Parse both files
    print(f"\nParsing CY{prev_year}...")
    parsed_prev = parse_post_processor_output(file_path=prev_file, year=prev_year)
    if not parsed_prev:
        print(f"ERROR: Failed to parse {prev_file}")
        return 1

    print(f"\nParsing CY{year}...")
    parsed_curr = parse_post_processor_output(file_path=curr_file, year=year)
    if not parsed_curr:
        print(f"ERROR: Failed to parse {curr_file}")
        return 1

    # Validate required streams exist in parsed data
    sample_year_prev = next(iter(parsed_prev))
    sample_year_curr = next(iter(parsed_curr))
    for s in STREAMS:
        if s not in parsed_prev.get(sample_year_prev, {}):
            print(f"ERROR: {s} not found in CY{prev_year} data (year {sample_year_prev})")
            print("  Post-processor may have failed or output format changed.")
            return 1
        if s not in parsed_curr.get(sample_year_curr, {}):
            print(f"ERROR: {s} not found in CY{year} data (year {sample_year_curr})")
            print("  Post-processor may have failed or output format changed.")
            return 1

    # Compute superposition series for each stream from each model run
    print("\nComputing superposition series...")
    data: dict[str, dict[str, dict[int, float]]] = {}
    for stream in STREAMS:
        print(f"\n  {STREAM_LABELS[stream]} ({stream}):")
        prev_series = compute_superposition_series(parsed_prev, stream)
        curr_series = compute_superposition_series(parsed_curr, stream)
        data[stream] = {"prev": prev_series, "curr": curr_series}
        print(f"    CY{prev_year}: {len(prev_series)} years computed")
        print(f"    CY{year}: {len(curr_series)} years computed")

    # Write xlsx
    output_dir = Path("./output/depletion/")
    output_dir.mkdir(parents=True, exist_ok=True)
    output_path = output_dir / f"Table_3_verify_depletion_{prev_year}_{year}.xlsx"
    write_verification_xlsx(output_path, prev_year, year, data)
    print(f"\nOutput written to: {output_path}")

    # Console summary
    print("\n" + "=" * 65)
    print("VERIFICATION SUMMARY")
    print("=" * 65)

    has_flags = False

    for stream in STREAMS:
        label = STREAM_LABELS[stream]
        prev_series = data[stream]["prev"]
        curr_series = data[stream]["curr"]

        print(f"\n  {label}:")

        # Historical years (should match)
        historical_diffs: list[tuple[int, float]] = []
        for yr in range(YEAR_START, year):
            if yr in prev_series and yr in curr_series:
                diff = abs(curr_series[yr] - prev_series[yr])
                historical_diffs.append((yr, diff))

        if historical_diffs:
            max_hist_yr, max_hist_diff = max(historical_diffs, key=lambda x: x[1])
            print(f"    Historical years ({YEAR_START}-{year - 1}):")
            print(f"      Max |diff|: {max_hist_diff:.6f} AF (year {max_hist_yr})")
            if max_hist_diff > DIFF_THRESHOLD_AF:
                print(f"      FLAG: Exceeds {DIFF_THRESHOLD_AF} AF threshold")
                has_flags = True
                # Print all flagged years
                flagged = [(yr, d) for yr, d in historical_diffs if d > DIFF_THRESHOLD_AF]
                for yr, d in flagged:
                    print(f"        Year {yr}: diff = {d:.6f} AF")
            else:
                print(f"      PASS: Within {DIFF_THRESHOLD_AF} AF threshold")

        # Current year and future (expected to differ)
        future_diffs: list[tuple[int, float]] = []
        for yr in range(year, YEAR_END + 1):
            if yr in prev_series and yr in curr_series:
                diff = curr_series[yr] - prev_series[yr]
                future_diffs.append((yr, diff))

        if future_diffs:
            print(f"    Current + future years ({year}-{YEAR_END}):")
            for yr, diff in future_diffs:
                status = "  <-- new pumping data" if yr == year else ""
                print(f"      Year {yr}: diff = {diff:+.6f} AF{status}")

    print("\n" + "=" * 65)
    if has_flags:
        print("VERDICT: FLAG — Historical year differences exceed threshold")
        print("         Review xlsx for details (yellow-highlighted cells)")
    else:
        print("VERDICT: PASS — Historical years consistent between model runs")
    print("=" * 65)

    return 1 if has_flags else 0


if __name__ == "__main__":
    parser = argparse.ArgumentParser(
        description="Verify depletion consistency between consecutive MODFLOW runs"
    )
    parser.add_argument(
        "--year",
        type=int,
        default=2025,
        help="Current processing year (compares against year-1). Default: 2025",
    )
    args = parser.parse_args()
    sys.exit(main(args.year))
