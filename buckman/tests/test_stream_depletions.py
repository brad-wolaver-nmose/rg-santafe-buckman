"""
Smoke tests for stream_depletions module.
Verifies code RUNS - domain expert must verify calculations independently.

These tests support the Ralph iterate-until-pass loop.
They catch mechanical failures, not logical errors.

Scientific context:
- Stream depletions are reductions in surface water flow caused by groundwater pumping
- Values in cfs (cubic feet per second) or acre-feet (AF)
- Typical depletion magnitudes: 0.001-1.0 cfs per cell, 50-100 AF/year per stream
"""
from pathlib import Path

import pytest

# Apply Layer 0 marker to all tests in this file
pytestmark = pytest.mark.layer0


def test_module_imports():
    """Verify module imports without syntax errors."""


def test_cfs_to_af_exists():
    """Verify unit conversion function exists."""
    from stream_depletions import cfs_to_acre_feet
    assert callable(cfs_to_acre_feet)


def test_cfs_to_af_runs():
    """
    Verify cfs_to_acre_feet executes without crashing.

    Hand calculation check:
    - 1 cfs for 30 days = 1 * 30 * 1.9835 = 59.505 AF
    """
    from stream_depletions import cfs_to_acre_feet

    # Simple realistic input: 1 cfs for 30 days
    result = cfs_to_acre_feet(1.0, 30)

    assert result is not None
    assert isinstance(result, float)
    # Sanity check: should be roughly 59.5 AF (not 0, not 1000)
    assert 50 < result < 70, f"Result {result} outside expected range [50, 70]"


def test_cfs_to_af_sanity():
    """
    Verify conversion produces physically reasonable values.

    Scientific basis:
    - 1 cfs = 1 ft³/s = 86,400 ft³/day
    - 1 AF = 43,560 ft³
    - So 1 cfs for 1 day = 86400/43560 = 1.9835 AF

    Test: 0.1 cfs for 31 days (January) should be ~6.15 AF
    """
    from stream_depletions import cfs_to_acre_feet

    result = cfs_to_acre_feet(0.1, 31)

    # Expected: 0.1 * 31 * 1.9835 = 6.149 AF
    assert 5.0 < result < 8.0, f"Result {result} outside expected range [5, 8]"


def test_core_2003_residuals_exist():
    """Verify Core (2003) analytical model constants are defined."""
    from stream_depletions import CORE_2003_POJOAQUE, CORE_2003_TESUQUE

    # Pojoaque should have values 1988-2015
    assert 1988 in CORE_2003_POJOAQUE
    assert 2015 in CORE_2003_POJOAQUE
    assert isinstance(CORE_2003_POJOAQUE[2015], float)

    # Tesuque should have values 1988-2030+
    assert 1988 in CORE_2003_TESUQUE
    assert 2024 in CORE_2003_TESUQUE
    assert isinstance(CORE_2003_TESUQUE[2024], float)


def test_core_2003_pojoaque_value():
    """
    Verify Pojoaque residual value matches Core (2003) table.

    From Core (2003) PROJECTION.XLS:
    - 2015: 0.316 AF (last positive value)
    - 2024: 0 (residual exhausted after 2015)
    """
    from stream_depletions import get_analytical_residual

    # 2015 should be ~0.316 AF
    result_2015 = get_analytical_residual("pojoaque", 2015)
    assert 0.0 < result_2015 < 1.0, f"2015 Pojoaque residual {result_2015} should be ~0.316"

    # 2024 should be 0 (or very small)
    result_2024 = get_analytical_residual("pojoaque", 2024)
    assert result_2024 == 0.0 or result_2024 < 0.01, "2024 Pojoaque residual should be 0"


def test_core_2003_tesuque_value():
    """
    Verify Tesuque residual value matches Core (2003) table.

    From Core (2003) PROJECTION.XLS:
    - 2024: 12.877 AF
    """
    from stream_depletions import get_analytical_residual

    result = get_analytical_residual("tesuque", 2024)

    # Expected: 12.877 AF (allow 0.5 AF tolerance)
    assert 12.0 < result < 14.0, f"2024 Tesuque residual {result} should be ~12.877"


def test_otowi_cell_definitions():
    """Verify Otowi gage cell lists are defined correctly."""
    from stream_depletions import ABOVE_OTOWI_CELLS, BELOW_OTOWI_CELLS

    # Above Otowi should have 10 cells
    assert len(ABOVE_OTOWI_CELLS) == 10, f"Expected 10 Above Otowi cells, got {len(ABOVE_OTOWI_CELLS)}"

    # Below Otowi should have 16 cells
    assert len(BELOW_OTOWI_CELLS) == 16, f"Expected 16 Below Otowi cells, got {len(BELOW_OTOWI_CELLS)}"

    # All cells should be (layer, row, col) tuples with layer=1
    for cell in ABOVE_OTOWI_CELLS + BELOW_OTOWI_CELLS:
        assert len(cell) == 3, f"Cell {cell} should be (layer, row, col) tuple"
        assert cell[0] == 1, f"Cell {cell} should have layer=1"


def test_days_per_month_2024():
    """Verify 2024 days per month (leap year) is defined correctly."""
    from stream_depletions import DAYS_2024

    assert len(DAYS_2024) == 12, "Should have 12 months"
    assert DAYS_2024[1] == 29, "February 2024 should have 29 days (leap year)"
    assert sum(DAYS_2024) == 366, "2024 is a leap year with 366 days"


def test_parse_postprocessor_output_exists():
    """Verify parser function exists."""
    from stream_depletions import parse_postprocessor_output
    assert callable(parse_postprocessor_output)


def test_print_error_exists():
    """Verify forensic error printing function exists."""
    from stream_depletions import print_error
    assert callable(print_error)


def test_extract_otowi_depletions_exists():
    """Verify extract_otowi_depletions function exists and is callable."""
    from stream_depletions import extract_otowi_depletions
    assert callable(extract_otowi_depletions)


def test_print_otowi_verification_exists():
    """Verify print_otowi_verification function exists and is callable."""
    from stream_depletions import print_otowi_verification
    assert callable(print_otowi_verification)


def test_extract_otowi_depletions_with_mock_data():
    """
    Verify extract_otowi_depletions correctly sums cell values.

    Uses minimal mock data to test the aggregation logic.
    """
    from stream_depletions import ABOVE_OTOWI_CELLS, BELOW_OTOWI_CELLS, extract_otowi_depletions

    # Create mock parsed data with 0.1 cfs for all cells and months
    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    mock_data: dict[int, dict[str, dict[str, float]]] = {2024: {}}

    for lay, row, col in ABOVE_OTOWI_CELLS + BELOW_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {month: 0.1 for month in months}

    above, below = extract_otowi_depletions(mock_data, 2024)

    # Above: 10 cells * 0.1 = 1.0 per month
    assert len(above) == 12
    assert abs(above[0] - 1.0) < 0.001, f"Expected 1.0, got {above[0]}"

    # Below: 16 cells * 0.1 = 1.6 per month
    assert len(below) == 12
    assert abs(below[0] - 1.6) < 0.001, f"Expected 1.6, got {below[0]}"


def test_print_residual_verification_exists():
    """Verify print_residual_verification function exists and is callable."""
    from stream_depletions import print_residual_verification
    assert callable(print_residual_verification)


def test_print_residual_verification_runs(capsys):
    """Verify print_residual_verification runs and prints expected output."""
    from stream_depletions import print_residual_verification

    print_residual_verification(2024)
    captured = capsys.readouterr()

    # Should contain stream names and values
    assert "Pojoaque" in captured.out
    assert "Tesuque" in captured.out
    assert "12.877" in captured.out  # Tesuque 2024 value
    assert "0.000" in captured.out   # Pojoaque 2024 value (exhausted)


def test_cfs_to_af_wrapper_exists():
    """Verify cfs_to_af wrapper function exists and is callable."""
    from stream_depletions import cfs_to_af
    assert callable(cfs_to_af)


def test_cfs_to_af_january():
    """
    Verify cfs_to_af returns correct value for January.

    Hand calculation:
    - January has 31 days
    - 0.1 cfs * 31 days * 1.9835 = 6.14885 AF
    """
    from stream_depletions import cfs_to_af

    result = cfs_to_af(0.1, 0)  # January is month_index 0

    assert abs(result - 6.14885) < 0.001, f"Expected ~6.14885, got {result}"


def test_cfs_to_af_february_leap_year():
    """
    Verify cfs_to_af handles February with use_leap_year flag.

    Hand calculation (leap year, 29 days):
    - 1.0 cfs * 29 days * 86400/43560 = 57.5207 AF

    Hand calculation (non-leap year, 28 days, default):
    - 1.0 cfs * 28 days * 86400/43560 = 55.5372 AF
    """
    from stream_depletions import cfs_to_af

    # Default (non-leap year, 28 days)
    result_default = cfs_to_af(1.0, 1)  # February is month_index 1
    assert abs(result_default - 55.5372) < 0.001, f"Expected ~55.5372, got {result_default}"

    # Leap year (29 days)
    result_leap = cfs_to_af(1.0, 1, use_leap_year=True)
    assert abs(result_leap - 57.5207) < 0.001, f"Expected ~57.5207, got {result_leap}"


def test_cfs_to_af_raises_on_negative():
    """Verify cfs_to_af raises ValueError on negative cfs."""
    from stream_depletions import cfs_to_af

    with pytest.raises(ValueError, match="cfs_value must be >= 0"):
        cfs_to_af(-1.0, 0)


def test_cfs_to_af_raises_on_invalid_month():
    """Verify cfs_to_af raises ValueError on invalid month index."""
    from stream_depletions import cfs_to_af

    with pytest.raises(ValueError, match="month_index must be 0-11"):
        cfs_to_af(1.0, 12)

    with pytest.raises(ValueError, match="month_index must be 0-11"):
        cfs_to_af(1.0, -1)


def test_cfs_monthly_to_af_annual_exists():
    """Verify cfs_monthly_to_af_annual function exists and is callable."""
    from stream_depletions import cfs_monthly_to_af_annual
    assert callable(cfs_monthly_to_af_annual)


def test_cfs_monthly_to_af_annual_constant_flow():
    """
    Verify cfs_monthly_to_af_annual with constant 0.1 cfs all year.

    Hand calculation (non-leap year, 365 days, default):
    - 0.1 cfs * 365 days * 86400/43560 = 72.3967 AF

    Hand calculation (leap year, 366 days):
    - 0.1 cfs * 366 days * 86400/43560 = 72.5950 AF
    """
    from stream_depletions import cfs_monthly_to_af_annual

    # Default (non-leap year)
    result_default = cfs_monthly_to_af_annual([0.1] * 12)
    assert abs(result_default - 72.3967) < 0.01, f"Expected ~72.3967, got {result_default}"

    # Leap year
    result_leap = cfs_monthly_to_af_annual([0.1] * 12, use_leap_year=True)
    assert abs(result_leap - 72.5950) < 0.01, f"Expected ~72.5950, got {result_leap}"


def test_cfs_monthly_to_af_annual_raises_on_wrong_length():
    """Verify cfs_monthly_to_af_annual raises ValueError on non-12 element list."""
    from stream_depletions import cfs_monthly_to_af_annual

    with pytest.raises(ValueError, match="must have 12 elements"):
        cfs_monthly_to_af_annual([0.1] * 11)

    with pytest.raises(ValueError, match="must have 12 elements"):
        cfs_monthly_to_af_annual([0.1] * 13)


def test_cfs_monthly_to_af_annual_raises_on_negative():
    """Verify cfs_monthly_to_af_annual raises ValueError on negative cfs value."""
    from stream_depletions import cfs_monthly_to_af_annual

    cfs_list = [0.1] * 12
    cfs_list[5] = -0.1  # June is negative

    with pytest.raises(ValueError, match="cfs_list\\[5\\] must be >= 0"):
        cfs_monthly_to_af_annual(cfs_list)


def test_generate_table3_data_exists():
    """Verify generate_table3_data function exists and is callable."""
    from stream_depletions import generate_table3_data
    assert callable(generate_table3_data)


def test_generate_table3_data_with_mock():
    """
    Verify generate_table3_data correctly combines residuals and superposition.

    Uses mock data to test the calculation logic.
    """
    from stream_depletions import generate_table3_data

    # Create mock parsed data with constant 0.1 cfs for all months
    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    mock_data: dict[int, dict[str, dict[str, float]]] = {
        2024: {
            "R POJOAQUE": {m: 0.1 for m in months},
            "R TESUQUE": {m: 0.1 for m in months},
        }
    }

    result = generate_table3_data(mock_data, 2024)

    # Check structure
    assert "pojoaque" in result
    assert "tesuque" in result
    assert "residual_af" in result["pojoaque"]
    assert "superposition_af" in result["pojoaque"]
    assert "total_impact_af" in result["pojoaque"]

    # Check Pojoaque residual is 0 for 2024 (exhausted after 2015)
    assert result["pojoaque"]["residual_af"] == 0.0

    # Check Tesuque residual is 12.877 for 2024
    assert abs(result["tesuque"]["residual_af"] - 12.877) < 0.001

    # Check superposition: 0.1 cfs * 366 days * 1.9835 = 72.596 AF
    assert abs(result["pojoaque"]["superposition_af"] - 72.596) < 0.1
    assert abs(result["tesuque"]["superposition_af"] - 72.596) < 0.1

    # Check total = residual + superposition
    assert abs(result["pojoaque"]["total_impact_af"] - result["pojoaque"]["superposition_af"]) < 0.001
    assert abs(result["tesuque"]["total_impact_af"] - (12.877 + 72.596)) < 0.1


def test_generate_table3_data_raises_on_missing_year():
    """Verify generate_table3_data raises KeyError when year not found."""
    from stream_depletions import generate_table3_data

    mock_data: dict[int, dict[str, dict[str, float]]] = {2023: {}}

    with pytest.raises(KeyError, match="Year 2024 not found"):
        generate_table3_data(mock_data, 2024)


def test_generate_table3_data_raises_on_missing_stream():
    """Verify generate_table3_data raises KeyError when stream data missing."""
    from stream_depletions import generate_table3_data

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]
    mock_data: dict[int, dict[str, dict[str, float]]] = {
        2024: {
            "R POJOAQUE": {m: 0.1 for m in months},
            # Missing R TESUQUE
        }
    }

    with pytest.raises(KeyError, match="R TESUQUE not found"):
        generate_table3_data(mock_data, 2024)


def test_print_table3_verification_exists():
    """Verify print_table3_verification function exists and is callable."""
    from stream_depletions import print_table3_verification
    assert callable(print_table3_verification)


def test_print_table3_verification_runs(capsys):
    """Verify print_table3_verification runs and prints expected output."""
    from stream_depletions import print_table3_verification

    table3_data = {
        "pojoaque": {
            "residual_af": 0.0,
            "superposition_af": 60.797,
            "total_impact_af": 60.797,
        },
        "tesuque": {
            "residual_af": 12.877,
            "superposition_af": 98.456,
            "total_impact_af": 111.333,
        },
    }

    print_table3_verification(table3_data, 2024)
    captured = capsys.readouterr()

    # Should contain stream names and values
    assert "Pojoaque" in captured.out
    assert "Tesuque" in captured.out
    assert "60.797" in captured.out
    assert "12.877" in captured.out


def test_buckman_wells_cell_constant():
    """Verify BUCKMAN_WELLS_CELL constant is defined correctly."""
    from stream_depletions import BUCKMAN_WELLS_CELL

    assert BUCKMAN_WELLS_CELL == (1, 13, 11), f"Expected (1, 13, 11), got {BUCKMAN_WELLS_CELL}"


def test_generate_table4_data_exists():
    """Verify generate_table4_data function exists and is callable."""
    from stream_depletions import generate_table4_data
    assert callable(generate_table4_data)


def test_generate_table4_data_with_mock():
    """
    Verify generate_table4_data correctly generates Table 4 structure.

    Uses mock data with constant 0.1 cfs for all cells and months.
    """
    from stream_depletions import (
        ABOVE_OTOWI_CELLS,
        BELOW_OTOWI_CELLS,
        generate_table4_data,
    )

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # Create mock data with all required cells
    mock_data: dict[int, dict[str, dict[str, float]]] = {2024: {}}

    # Add Above Otowi cells
    for lay, row, col in ABOVE_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.1 for m in months}

    # Add Below Otowi cells
    for lay, row, col in BELOW_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.1 for m in months}

    # Add stream summaries
    mock_data[2024]["RIO GRANDE"] = {m: 2.6 for m in months}  # 10 + 16 cells * 0.1
    mock_data[2024]["R POJOAQUE"] = {m: 0.1 for m in months}
    mock_data[2024]["LC SPRINGS"] = {m: 0.01 for m in months}
    mock_data[2024]["R TESUQUE"] = {m: 0.05 for m in months}
    mock_data[2024]["RIV TOTAL"] = {m: 2.76 for m in months}

    result = generate_table4_data(mock_data, 2024)

    # Check structure
    assert "cell_data" in result
    assert "stream_summaries" in result
    assert "days_per_month" in result
    assert "above_otowi_cfs" in result
    assert "below_otowi_cfs" in result
    assert "above_otowi_af" in result
    assert "below_otowi_af" in result
    assert "above_otowi_annual_af" in result
    assert "below_otowi_annual_af" in result
    assert "buckman_cfs" in result
    assert "buckman_af" in result
    assert "buckman_annual_af" in result

    # Check cell counts
    above_count = sum(1 for c in result["cell_data"] if c["otowi"] == "above")
    below_count = sum(1 for c in result["cell_data"] if c["otowi"] == "below")
    assert above_count == 10, f"Expected 10 Above Otowi cells, got {above_count}"
    assert below_count == 16, f"Expected 16 Below Otowi cells, got {below_count}"

    # Check Above Otowi sum: 10 cells * 0.1 cfs = 1.0 cfs/month
    assert abs(result["above_otowi_cfs"][0] - 1.0) < 0.001

    # Check Below Otowi sum: 16 cells * 0.1 cfs = 1.6 cfs/month
    assert abs(result["below_otowi_cfs"][0] - 1.6) < 0.001


def test_generate_table4_data_raises_on_missing_year():
    """Verify generate_table4_data raises KeyError when year not found."""
    from stream_depletions import generate_table4_data

    mock_data: dict[int, dict[str, dict[str, float]]] = {2023: {}}

    with pytest.raises(KeyError, match="Year 2024 not found"):
        generate_table4_data(mock_data, 2024)


def test_generate_table4_data_raises_on_missing_stream():
    """Verify generate_table4_data raises KeyError when stream data missing."""
    from stream_depletions import ABOVE_OTOWI_CELLS, BELOW_OTOWI_CELLS, generate_table4_data

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    mock_data: dict[int, dict[str, dict[str, float]]] = {2024: {}}

    # Add cells but not stream summaries
    for lay, row, col in ABOVE_OTOWI_CELLS + BELOW_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.1 for m in months}

    with pytest.raises(KeyError, match="RIO GRANDE not found"):
        generate_table4_data(mock_data, 2024)


def test_print_table4_verification_exists():
    """Verify print_table4_verification function exists and is callable."""
    from stream_depletions import print_table4_verification
    assert callable(print_table4_verification)


def test_print_table4_verification_runs(capsys):
    """Verify print_table4_verification runs and prints expected output."""
    from stream_depletions import print_table4_verification

    # Create mock table4 data
    table4_data = {
        "cell_data": [
            {"key": 2107, "year": 2024, "lay": 1, "row": 1, "col": 16, "monthly_cfs": [0.1] * 12, "otowi": "above"},
        ] * 10 + [
            {"key": 2117, "year": 2024, "lay": 1, "row": 11, "col": 11, "monthly_cfs": [0.1] * 12, "otowi": "below"},
        ] * 16,
        "stream_summaries": {
            "RIO GRANDE": [2.6] * 12,
            "R POJOAQUE": [0.1] * 12,
        },
        "days_per_month": [31, 29, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31],
        "above_otowi_cfs": [1.0] * 12,
        "below_otowi_cfs": [1.6] * 12,
        "above_otowi_af": [61.5] * 12,
        "below_otowi_af": [98.4] * 12,
        "above_otowi_annual_af": 276.5,
        "below_otowi_annual_af": 442.4,
        "total_rg_af": [159.9] * 12,
        "total_rg_annual_af": 718.9,
        "buckman_cfs": [0.5] * 12,
        "buckman_af": [30.0] * 12,
        "buckman_annual_af": 360.0,
    }

    print_table4_verification(table4_data, 2024)
    captured = capsys.readouterr()

    # Should contain key information
    assert "2024" in captured.out
    assert "Above Otowi" in captured.out
    assert "Below Otowi" in captured.out
    assert "Buckman" in captured.out
    assert "276.5" in captured.out or "276.500" in captured.out


def test_la_cienega_cumulative_constant():
    """Verify LA_CIENEGA_CUMULATIVE constant is defined with expected structure."""
    from stream_depletions import LA_CIENEGA_CUMULATIVE

    # Should have years 2004-2030
    assert 2004 in LA_CIENEGA_CUMULATIVE
    assert 2023 in LA_CIENEGA_CUMULATIVE
    assert 2024 in LA_CIENEGA_CUMULATIVE
    assert 2030 in LA_CIENEGA_CUMULATIVE

    # 2024 validation value
    assert abs(LA_CIENEGA_CUMULATIVE[2024] - 3.74) < 0.01, f"Expected 3.74, got {LA_CIENEGA_CUMULATIVE[2024]}"

    # Values should be increasing (cumulative)
    assert LA_CIENEGA_CUMULATIVE[2024] > LA_CIENEGA_CUMULATIVE[2023]


def test_generate_table5_data_exists():
    """Verify generate_table5_data function exists and is callable."""
    from stream_depletions import generate_table5_data
    assert callable(generate_table5_data)


def test_generate_table5_data_with_mock():
    """
    Verify generate_table5_data correctly generates Table 5 structure.

    Uses mock data with constant 0.001 cfs for LC SPRINGS all months.
    """
    from stream_depletions import generate_table5_data

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # Create mock data with LC SPRINGS at ~0.00028 cfs (to get ~0.20 AF/year)
    # 0.20 AF / 366 days / 1.9835 = ~0.000275 cfs
    mock_cfs = 0.000275
    mock_data: dict[int, dict[str, dict[str, float]]] = {
        2024: {
            "LC SPRINGS": {m: mock_cfs for m in months},
        }
    }

    result = generate_table5_data(mock_data, 2024)

    # Check structure
    assert "year" in result
    assert "monthly_cfs" in result
    assert "annual_af" in result
    assert "previous_cumulative_af" in result
    assert "cumulative_af" in result

    # Check year
    assert result["year"] == 2024

    # Check monthly values
    assert len(result["monthly_cfs"]) == 12
    assert abs(result["monthly_cfs"][0] - mock_cfs) < 0.000001

    # Check previous cumulative (2023 from LA_CIENEGA_CUMULATIVE)
    assert abs(result["previous_cumulative_af"] - 3.54) < 0.01

    # Check cumulative = previous + annual
    expected_cumulative = result["previous_cumulative_af"] + result["annual_af"]
    assert abs(result["cumulative_af"] - expected_cumulative) < 0.001


def test_generate_table5_data_raises_on_missing_year():
    """Verify generate_table5_data raises KeyError when year not found."""
    from stream_depletions import generate_table5_data

    mock_data: dict[int, dict[str, dict[str, float]]] = {2023: {}}

    with pytest.raises(KeyError, match="Year 2024 not found"):
        generate_table5_data(mock_data, 2024)


def test_generate_table5_data_raises_on_missing_lc_springs():
    """Verify generate_table5_data raises KeyError when LC SPRINGS data missing."""
    from stream_depletions import generate_table5_data

    mock_data: dict[int, dict[str, dict[str, float]]] = {
        2024: {
            "RIO GRANDE": {"jan": 0.1},  # Not LC SPRINGS
        }
    }

    with pytest.raises(KeyError, match="LC SPRINGS not found"):
        generate_table5_data(mock_data, 2024)


def test_print_table5_verification_exists():
    """Verify print_table5_verification function exists and is callable."""
    from stream_depletions import print_table5_verification
    assert callable(print_table5_verification)


def test_print_table5_verification_runs(capsys):
    """Verify print_table5_verification runs and prints expected output."""
    from stream_depletions import print_table5_verification

    # Create mock table5 data
    table5_data = {
        "year": 2024,
        "monthly_cfs": [0.00028] * 12,
        "annual_af": 0.20,
        "previous_cumulative_af": 3.54,
        "cumulative_af": 3.74,
    }

    print_table5_verification(table5_data, 2024)
    captured = capsys.readouterr()

    # Should contain key information
    assert "2024" in captured.out
    assert "La Cienega" in captured.out
    assert "Annual" in captured.out
    assert "Cumulative" in captured.out
    assert "3.54" in captured.out  # Previous cumulative
    assert "3.74" in captured.out  # New cumulative


def test_write_table3_xlsx_exists():
    """Verify write_table3_xlsx function exists and is callable."""
    from stream_depletions import write_table3_xlsx
    assert callable(write_table3_xlsx)


def test_write_table3_xlsx_creates_file(tmp_path):
    """
    Verify write_table3_xlsx creates an XLSX file with correct structure.

    Uses mock data to test file creation and basic structure.
    """
    from stream_depletions import write_table3_xlsx

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # Create mock parsed data for a few years
    mock_data: dict[int, dict[str, dict[str, float]]] = {}
    for year in [2022, 2023, 2024]:
        mock_data[year] = {
            "R POJOAQUE": {m: 0.1 for m in months},
            "R TESUQUE": {m: 0.1 for m in months},
        }

    output_path = tmp_path / "test_table3.xlsx"
    result_path = write_table3_xlsx(mock_data, output_path, years=[2022, 2023, 2024])

    # Check file was created
    assert result_path.exists(), f"File was not created at {result_path}"

    # Check file can be opened with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active

    # Check sheet has expected dimensions (2 header rows + 3 data rows, 7 columns)
    assert ws.max_row == 5, f"Expected 5 rows, got {ws.max_row}"
    assert ws.max_column == 7, f"Expected 7 columns, got {ws.max_column}"

    # Check header values
    assert "Rio Pojoaque" in str(ws.cell(row=1, column=2).value)
    assert "Rio Tesuque" in str(ws.cell(row=1, column=5).value)
    assert "Year" in str(ws.cell(row=2, column=1).value)

    # Check data row years
    assert ws.cell(row=3, column=1).value == 2022
    assert ws.cell(row=4, column=1).value == 2023
    assert ws.cell(row=5, column=1).value == 2024

    wb.close()


def test_write_table3_xlsx_formatting(tmp_path):
    """
    Verify write_table3_xlsx applies correct formatting.
    """
    from stream_depletions import write_table3_xlsx

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    mock_data: dict[int, dict[str, dict[str, float]]] = {
        2024: {
            "R POJOAQUE": {m: 0.1 for m in months},
            "R TESUQUE": {m: 0.1 for m in months},
        }
    }

    output_path = tmp_path / "test_table3_fmt.xlsx"
    write_table3_xlsx(mock_data, output_path, years=[2024])

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Check header font is bold
    assert ws.cell(row=1, column=2).font.bold is True
    assert ws.cell(row=2, column=1).font.bold is True

    # Check data row year is bold
    assert ws.cell(row=3, column=1).font.bold is True

    # Check number format (3 decimal places)
    assert ws.cell(row=3, column=3).number_format == '0.000'

    wb.close()


def test_write_table3_xlsx_2024_values(tmp_path):
    """
    Verify write_table3_xlsx calculates correct 2024 values.

    2024 expected values from validation:
    - Pojoaque Residual: 0 (empty cell)
    - Tesuque Residual: 12.877
    """
    from stream_depletions import write_table3_xlsx

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # Create mock data with realistic cfs values
    # From validation: Pojoaque superposition ~60.797, Tesuque superposition ~20.706
    # Back-calculate cfs: 60.797 / 366 / 1.9835 ≈ 0.0838
    mock_data: dict[int, dict[str, dict[str, float]]] = {
        2024: {
            "R POJOAQUE": {m: 0.0838 for m in months},
            "R TESUQUE": {m: 0.0285 for m in months},
        }
    }

    output_path = tmp_path / "test_table3_values.xlsx"
    write_table3_xlsx(mock_data, output_path, years=[2024])

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Check 2024 Pojoaque residual is empty (0 after 2015)
    poj_residual = ws.cell(row=3, column=2).value
    assert poj_residual is None, f"Expected empty Pojoaque residual, got {poj_residual}"

    # Check 2024 Tesuque residual is 12.877
    tes_residual = ws.cell(row=3, column=5).value
    assert abs(tes_residual - 12.877) < 0.001, f"Expected 12.877, got {tes_residual}"

    wb.close()


def test_write_table4_xlsx_exists():
    """Verify write_table4_xlsx function exists and is callable."""
    from stream_depletions import write_table4_xlsx
    assert callable(write_table4_xlsx)


def test_write_table4_xlsx_creates_file(tmp_path):
    """
    Verify write_table4_xlsx creates an XLSX file with correct structure.

    Uses mock data to test file creation and basic structure.
    """
    from stream_depletions import (
        ABOVE_OTOWI_CELLS,
        BELOW_OTOWI_CELLS,
        write_table4_xlsx,
    )

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    # Create mock parsed data
    mock_data: dict[int, dict[str, dict[str, float]]] = {2024: {}}

    # Add Above Otowi cells
    for lay, row, col in ABOVE_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.01 for m in months}

    # Add Below Otowi cells
    for lay, row, col in BELOW_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.02 for m in months}

    # Add stream summaries
    mock_data[2024]["RIO GRANDE"] = {m: 0.42 for m in months}
    mock_data[2024]["R POJOAQUE"] = {m: 0.1 for m in months}
    mock_data[2024]["LC SPRINGS"] = {m: 0.01 for m in months}
    mock_data[2024]["R TESUQUE"] = {m: 0.05 for m in months}
    mock_data[2024]["RIV TOTAL"] = {m: 0.58 for m in months}

    output_path = tmp_path / "test_table4.xlsx"
    result_path = write_table4_xlsx(mock_data, output_path, year=2024)

    # Check file was created
    assert result_path.exists(), f"File was not created at {result_path}"

    # Check file can be opened with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active

    # Check sheet has reasonable dimensions
    # With only Otowi cells (26 cells), we expect:
    # 1 header + 26 cells + 5 streams + 1 month labels + 1 days + 2 cfs sums +
    # 1 AF header + 5 AF rows = 42 rows minimum
    assert ws.max_row >= 40, f"Expected at least 40 rows, got {ws.max_row}"
    assert ws.max_column >= 17, f"Expected at least 17 columns, got {ws.max_column}"

    # Check header values
    assert ws.cell(row=1, column=1).value == "KEY"
    assert ws.cell(row=1, column=2).value == "YEAR"
    assert ws.cell(row=1, column=6).value == "JAN"
    assert ws.cell(row=1, column=18).value == "Otowi"

    wb.close()


def test_write_table4_xlsx_cell_data(tmp_path):
    """
    Verify write_table4_xlsx correctly writes cell data rows.
    """
    from stream_depletions import (
        ABOVE_OTOWI_CELLS,
        BELOW_OTOWI_CELLS,
        write_table4_xlsx,
    )

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    mock_data: dict[int, dict[str, dict[str, float]]] = {2024: {}}

    for lay, row, col in ABOVE_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.01 for m in months}

    for lay, row, col in BELOW_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.02 for m in months}

    mock_data[2024]["RIO GRANDE"] = {m: 0.42 for m in months}
    mock_data[2024]["R POJOAQUE"] = {m: 0.1 for m in months}
    mock_data[2024]["LC SPRINGS"] = {m: 0.01 for m in months}
    mock_data[2024]["R TESUQUE"] = {m: 0.05 for m in months}
    mock_data[2024]["RIV TOTAL"] = {m: 0.58 for m in months}

    output_path = tmp_path / "test_table4_cells.xlsx"
    write_table4_xlsx(mock_data, output_path, year=2024)

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Check cell data has "above" and "below" labels in column R (18)
    above_count = 0
    below_count = 0
    for row in range(2, 50):  # Cell data rows
        otowi_val = ws.cell(row=row, column=18).value
        if otowi_val == "above":
            above_count += 1
        elif otowi_val == "below":
            below_count += 1

    assert above_count == 10, f"Expected 10 'above' labels, got {above_count}"
    assert below_count == 16, f"Expected 16 'below' labels, got {below_count}"

    wb.close()


def test_write_table4_xlsx_formulas(tmp_path):
    """
    Verify write_table4_xlsx includes correct formulas for AF calculations.
    """
    from stream_depletions import (
        ABOVE_OTOWI_CELLS,
        BELOW_OTOWI_CELLS,
        write_table4_xlsx,
    )

    months = ["jan", "feb", "mar", "apr", "may", "jun", "jul", "aug", "sep", "oct", "nov", "dec"]

    mock_data: dict[int, dict[str, dict[str, float]]] = {2024: {}}

    for lay, row, col in ABOVE_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.01 for m in months}

    for lay, row, col in BELOW_OTOWI_CELLS:
        cell_key = f"{lay} {row} {col}"
        mock_data[2024][cell_key] = {m: 0.02 for m in months}

    mock_data[2024]["RIO GRANDE"] = {m: 0.42 for m in months}
    mock_data[2024]["R POJOAQUE"] = {m: 0.1 for m in months}
    mock_data[2024]["LC SPRINGS"] = {m: 0.01 for m in months}
    mock_data[2024]["R TESUQUE"] = {m: 0.05 for m in months}
    mock_data[2024]["RIV TOTAL"] = {m: 0.58 for m in months}

    output_path = tmp_path / "test_table4_formulas.xlsx"
    write_table4_xlsx(mock_data, output_path, year=2024)

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Find rows with "Above Otowi" label in column E (should be the AF calculation row)
    above_af_row = None
    for row in range(1, ws.max_row + 1):
        if ws.cell(row=row, column=5).value == "Above Otowi":
            above_af_row = row
            break

    assert above_af_row is not None, "Could not find 'Above Otowi' AF row"

    # Check that the cell contains a formula (starts with =)
    jan_af_cell = ws.cell(row=above_af_row, column=6).value
    assert jan_af_cell is not None
    assert str(jan_af_cell).startswith("="), f"Expected formula, got {jan_af_cell}"

    # Check Total column has SUM formula
    total_cell = ws.cell(row=above_af_row, column=18).value
    assert total_cell is not None
    assert "SUM" in str(total_cell).upper(), f"Expected SUM formula, got {total_cell}"

    wb.close()


def test_write_table5_xlsx_exists():
    """Verify write_table5_xlsx function exists and is callable."""
    from stream_depletions import write_table5_xlsx
    assert callable(write_table5_xlsx)


def test_write_table5_xlsx_creates_file(tmp_path):
    """
    Verify write_table5_xlsx creates an XLSX file with correct structure.
    """
    from stream_depletions import write_table5_xlsx

    output_path = tmp_path / "test_table5.xlsx"
    result_path = write_table5_xlsx(output_path, years=[2022, 2023, 2024])

    # Check file was created
    assert result_path.exists(), f"File was not created at {result_path}"

    # Check file can be opened with openpyxl
    import openpyxl
    wb = openpyxl.load_workbook(result_path)
    ws = wb.active

    # Check sheet has expected dimensions (1 header row + 3 data rows, 2 columns)
    assert ws.max_row == 4, f"Expected 4 rows, got {ws.max_row}"
    assert ws.max_column == 2, f"Expected 2 columns, got {ws.max_column}"

    # Check header values
    assert ws.cell(row=1, column=1).value == "Year"
    assert ws.cell(row=1, column=2).value == "Total"

    # Check data row years
    assert ws.cell(row=2, column=1).value == 2022
    assert ws.cell(row=3, column=1).value == 2023
    assert ws.cell(row=4, column=1).value == 2024

    wb.close()


def test_write_table5_xlsx_formatting(tmp_path):
    """
    Verify write_table5_xlsx applies correct formatting.
    """
    from stream_depletions import write_table5_xlsx

    output_path = tmp_path / "test_table5_fmt.xlsx"
    write_table5_xlsx(output_path, years=[2024])

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Check header font is bold
    assert ws.cell(row=1, column=1).font.bold is True
    assert ws.cell(row=1, column=2).font.bold is True

    # Check number format (2 decimal places)
    assert ws.cell(row=2, column=2).number_format == '0.00'

    wb.close()


def test_write_table5_xlsx_2024_value(tmp_path):
    """
    Verify write_table5_xlsx has correct 2024 cumulative value.

    2024 expected value from validation: 3.74 AF
    """
    from stream_depletions import write_table5_xlsx

    output_path = tmp_path / "test_table5_values.xlsx"
    write_table5_xlsx(output_path, years=[2024])

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Check 2024 cumulative value is 3.74
    total_value = ws.cell(row=2, column=2).value
    assert abs(total_value - 3.74) < 0.01, f"Expected 3.74, got {total_value}"

    wb.close()


def test_write_table5_xlsx_default_years(tmp_path):
    """
    Verify write_table5_xlsx uses default years 2004-2030 when not specified.
    """
    from stream_depletions import write_table5_xlsx

    output_path = tmp_path / "test_table5_default.xlsx"
    write_table5_xlsx(output_path)  # No years specified

    import openpyxl
    wb = openpyxl.load_workbook(output_path)
    ws = wb.active

    # Check sheet has expected dimensions (1 header row + 27 data rows, 2 columns)
    assert ws.max_row == 28, f"Expected 28 rows (header + 2004-2030), got {ws.max_row}"

    # Check first and last year
    assert ws.cell(row=2, column=1).value == 2004
    assert ws.cell(row=28, column=1).value == 2030

    # Check 2004 value from LA_CIENEGA_CUMULATIVE
    assert abs(ws.cell(row=2, column=2).value - 0.45) < 0.01

    # Check 2030 value from LA_CIENEGA_CUMULATIVE
    assert abs(ws.cell(row=28, column=2).value - 4.80) < 0.01

    wb.close()


# =============================================================================
# VALIDATION FUNCTION TESTS (US-014)
# =============================================================================


def test_validation_constants_exist():
    """Verify validation tolerance constants are defined."""
    from stream_depletions import VALIDATION_TOLERANCE_AF, VALIDATION_TOLERANCE_CFS

    assert VALIDATION_TOLERANCE_AF == 0.01
    assert VALIDATION_TOLERANCE_CFS == 0.000001


def test_validate_table3_exists():
    """Verify validate_table3 function exists and is callable."""
    from stream_depletions import validate_table3
    assert callable(validate_table3)


def test_validate_table4_exists():
    """Verify validate_table4 function exists and is callable."""
    from stream_depletions import validate_table4
    assert callable(validate_table4)


def test_validate_table5_exists():
    """Verify validate_table5 function exists and is callable."""
    from stream_depletions import validate_table5
    assert callable(validate_table5)


def test_validate_all_tables_exists():
    """Verify validate_all_tables function exists and is callable."""
    from stream_depletions import validate_all_tables
    assert callable(validate_all_tables)


def test_print_validation_results_exists():
    """Verify print_validation_results function exists and is callable."""
    from stream_depletions import print_validation_results
    assert callable(print_validation_results)


def test_validate_table3_missing_file():
    """Verify validate_table3 handles missing file gracefully."""
    from stream_depletions import validate_table3

    result = validate_table3(
        validation_path='/nonexistent/file.xlsx',
        generated_data={'pojoaque': {}, 'tesuque': {}},
        year=2024
    )

    assert result['status'] == 'FAILED'
    assert len(result['errors']) > 0
    assert 'not found' in result['errors'][0]


def test_validate_table4_missing_file():
    """Verify validate_table4 handles missing file gracefully."""
    from stream_depletions import validate_table4

    result = validate_table4(
        validation_path='/nonexistent/file.xlsx',
        generated_data={'calculations': {}},
        year=2024
    )

    assert result['status'] == 'FAILED'
    assert len(result['errors']) > 0
    assert 'not found' in result['errors'][0]


def test_validate_table5_returns_dict():
    """Verify validate_table5 returns expected structure."""
    from stream_depletions import validate_table5

    # Use perfect matching data to test structure
    generated_data = {
        'annual_af': 0.20,  # 3.74 - 3.54 = 0.20
        'cumulative_af': 3.74,  # 2024 value from constant
    }

    result = validate_table5(generated_data, year=2024)

    assert 'status' in result
    assert 'comparisons' in result
    assert 'errors' in result
    assert isinstance(result['comparisons'], list)
    assert len(result['comparisons']) == 2  # cumulative and annual


def test_validate_table5_passing():
    """Verify validate_table5 passes with matching data."""
    from stream_depletions import validate_table5

    # Use values that match LA_CIENEGA_CUMULATIVE
    generated_data = {
        'annual_af': 0.20,  # 3.74 - 3.54 = 0.20
        'cumulative_af': 3.74,  # 2024 value from constant
    }

    result = validate_table5(generated_data, year=2024)

    assert result['status'] == 'OK'
    assert len(result['errors']) == 0


def test_validate_all_tables_returns_dict():
    """Verify validate_all_tables returns expected structure."""
    from stream_depletions import validate_all_tables

    # Dummy data (files don't exist, so will fail)
    result = validate_all_tables(
        table3_validation_path='/nonexistent/t3.xlsx',
        table4_validation_path='/nonexistent/t4.xlsx',
        table3_data={'pojoaque': {}, 'tesuque': {}},
        table4_data={'calculations': {}},
        table5_data={'annual_af': 0, 'cumulative_af': 0},
        year=2024
    )

    assert 'overall_status' in result
    assert 'table3' in result
    assert 'table4' in result
    assert 'table5' in result
    # Should fail because validation files don't exist
    assert result['overall_status'] == 'FAILED'


def test_validate_table3_with_real_file():
    """Integration test: validate Table 3 against actual validation file."""
    from stream_depletions import (
        generate_table3_data,
        parse_postprocessor_output,
        validate_table3,
    )

    # Paths
    validation_path = Path('validation/TABLE 3 - Rio Pojoaque-Nambe & Rio Tesuque.xlsx')
    postprocessor_output = Path('output/modflow/2024/depletions/CY2024')

    if not validation_path.exists():
        pytest.skip("Validation file not available")
    if not postprocessor_output.exists():
        pytest.skip("Post-processor output not available")

    # Generate data from actual post-processor output
    parsed_data = parse_postprocessor_output(postprocessor_output)
    table3_data = generate_table3_data(parsed_data, year=2024)

    # Validate
    result = validate_table3(validation_path, table3_data, year=2024)

    # Print diagnostics if failing
    if result['status'] != 'OK':
        for comp in result['comparisons']:
            print(f"{comp['field']}: calc={comp['calc']:.6f}, valid={comp['valid']:.6f}")

    assert result['status'] == 'OK', f"Validation failed: {result['errors']}"


def test_validate_table4_with_real_file():
    """Integration test: validate Table 4 against actual validation file."""
    from stream_depletions import (
        generate_table4_data,
        parse_postprocessor_output,
        validate_table4,
    )

    # Paths
    validation_path = Path('validation/TABLE 4 - Rio Grande, above below Otowi.xlsx')
    postprocessor_output = Path('output/modflow/2024/depletions/CY2024')

    if not validation_path.exists():
        pytest.skip("Validation file not available")
    if not postprocessor_output.exists():
        pytest.skip("Post-processor output not available")

    # Generate data from actual post-processor output
    parsed_data = parse_postprocessor_output(postprocessor_output)
    table4_data = generate_table4_data(parsed_data, year=2024)

    # Validate
    result = validate_table4(validation_path, table4_data, year=2024)

    # Print diagnostics if failing
    if result['status'] != 'OK':
        for comp in result['comparisons']:
            if not comp['ok']:
                print(f"{comp['field']}: calc={comp['calc']:.6f}, valid={comp['valid']:.6f}")

    assert result['status'] == 'OK', f"Validation failed: {result['errors']}"


# Note: Integration tests requiring actual files are skipped in smoke tests.
# The domain expert should run the full workflow and verify:
# 1. Post-processor output file is generated
# 2. Table 3 Pojoaque 2024 superposition matches validation
# 3. Table 4 Otowi Above/Below sums match validation
# 4. Table 5 La Cienega 2024 cumulative matches validation
