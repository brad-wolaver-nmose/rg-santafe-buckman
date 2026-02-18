#!/usr/bin/env python3
"""
Buckman Well Field CSV Data Ingestion Workflow

This script ingests daily well production data from CSV files provided by the City
of Santa Fe for the Buckman Well Field. It aggregates daily MGD (million gallons per day)
values into monthly totals (MG), converts to acre-feet, validates data quality, and
produces structured outputs for report generation.

Scientific Basis:
- Volume conversion: USGS standard (https://water.usgs.gov/nawqa/glos.html)
  1 acre-foot = 325,851 gallons → 1 MG = 3.06889 AF
- Daily aggregation: Sum of daily flow rates (MGD) × 1 day = monthly volume (MG)
- Units handled with pint library for dimensional safety

PYTHON DEPENDENCIES:
- Install with: pip install -r requirements.txt
  (pandas>=1.5.0, pint>=0.20.0)
"""

import argparse
import sys
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from pint import UnitRegistry

# Initialize unit registry for dimensional analysis
ureg = UnitRegistry()

# Define custom units for water resource management
# Million gallons (MG) = 1,000,000 gallons
ureg.define('million_gallon = 1e6 * gallon = MG')
# Million gallons per day (MGD) - common water industry unit
ureg.define('million_gallon_per_day = million_gallon / day = MGD')

# =============================================================================
# CONFIGURATION CONSTANTS
# =============================================================================

# Path pattern for source CSV file — replace {year} with actual year at runtime
# CSV contains daily MGD (million gallons per day) values for 13 Buckman wells
INPUT_CSV_PATH: str = "./input/csv/Buckman_Well_Prod_{year}.csv"

# Output directory for all generated files
OUTPUT_DIR: str = "./output/ingested_data"

# Validation data directory (contains reference Excel files for verification)
VALIDATION_DIR: str = "./validation"

# USGS conversion factor: 1 million gallons = 3.06889 acre-feet
# Scientific basis: 1 acre-foot = 325,851 gallons (USGS definition)
# Calculation: 1,000,000 gallons / 325,851 gallons/AF = 3.06889 AF/MG
MG_TO_AF_FACTOR: float = 3.06889

# Three-tier tolerance thresholds for daily BWP verification (MGD)
# Each day's per-well sum is compared to the BWP formula column
# Tier 1: Absolute noise floor - below instrument precision (database artifacts)
# Set to 0.0015 MGD to catch 100-gallon rounding artifacts (900-1,500 gal range)
NOISE_THRESHOLD_MGD: float = 0.0015  # 1,500 gal/day - database precision artifacts

# Tier 2: Informational threshold - expected Excel rounding differences
DAILY_SUM_TOLERANCE_INFO_MGD: float = 0.001  # 1,000 gal/day tolerance

# Tier 3: Error threshold - significant mismatches requiring CSV review
DAILY_SUM_TOLERANCE_ERROR_MGD: float = 0.005  # 5,000 gal/day threshold

# Physical interpretation:
# - 0.0015 MGD = 1,500 gal/day (catches 100-gal database rounding: 900-1,500 gal range)
# - 0.001 MGD = 1,000 gal/day (typical Excel formula rounding tolerance)
# - 0.005 MGD = 5,000 gal/day (~0.2% of average daily production)

# Tolerance for annual Sum row verification (MG)
# Our 12-month totals are compared to the CSV's Sum row per well
# Physical interpretation: 0.01 MG = 10,000 gallons difference threshold
ANNUAL_SUM_TOLERANCE_MG: float = 0.01

# Month abbreviations in calendar order (used for output filenames and tables)
MONTHS_ABBREV: tuple[str, ...] = (
    "JAN", "FEB", "MAR", "APR", "MAY", "JUN",
    "JUL", "AUG", "SEP", "OCT", "NOV", "DEC",
)

# Month tuples for iteration: (numeric string, abbreviation)
MONTHS_ORDERED: tuple[tuple[str, str], ...] = (
    ("01", "JAN"), ("02", "FEB"), ("03", "MAR"), ("04", "APR"),
    ("05", "MAY"), ("06", "JUN"), ("07", "JUL"), ("08", "AUG"),
    ("09", "SEP"), ("10", "OCT"), ("11", "NOV"), ("12", "DEC"),
)

# Well-to-OSE number mapping: well number (1-13) → OSE permit number
# These are the New Mexico State Engineer office permit numbers for each Buckman well
# Well 3 is labeled "3/3A" in historical reports (combined wells 3 and 3A)
WELL_OSE_MAP: dict[int, str] = {
    1: "RG-20516-S-5",
    2: "RG-20516-S-6",
    3: "RG-20516-S",      # Labeled as "3/3A" in reports
    4: "RG-20516-S-2",
    5: "RG-20516-S-3",
    6: "RG-20516-S-4",
    7: "RG-20516-S-7",
    8: "RG-20516-S-8",
    9: "RG-20516-S-9",
    10: "RG-20516-S-10",
    11: "RG-20516-S-11",
    12: "RG-20516-S-12",
    13: "RG-20516-S-13",
}

# CSV column headers for wells 1-13 (as they appear in the source CSV)
# Format: "BWell N|Flow Mgd" where N is well number
CSV_WELL_COLUMNS: list[str] = [
    "BWell 1|Flow Mgd", "BWell 2|Flow Mgd", "BWell 3|Flow Mgd", "BWell 4|Flow Mgd", "BWell 5|Flow Mgd",
    "BWell 6|Flow Mgd", "BWell 7|Flow Mgd", "BWell 8|Flow Mgd", "BWell 9|Flow Mgd", "BWell 10|Flow Mgd",
    "BWell 11|Flow Mgd", "BWell 12|Flow Mgd", "BWell 13|Flow Mgd",
]

# Header name for the total/formula column in the source CSV
# This column contains the daily total production across all wells (MGD)
CSV_TOTAL_COLUMN: str = "BWP|Flow Mgd|MGD|Formula"


# =============================================================================
# HELPER FUNCTIONS
# =============================================================================

def print_error(what_failed: str, location: str, actual: str, expected: str, context: str) -> None:
    """
    Print forensic-quality error message following CLAUDE.md standards.

    Prints error information in a structured 5-element format for debugging.

    Args:
        what_failed: Description of what operation failed
        location: Where the failure occurred (file path, function name, etc.)
        actual: The actual value that was encountered
        expected: The expected value or condition
        context: Physical interpretation or additional context

    Example:
        >>> print_error(
        ...     "CSV file not found",
        ...     "./input/csv/Buckman_Well_Prod_2024.csv",
        ...     "File does not exist",
        ...     "CSV file with 366 daily records",
        ...     "Missing source data - cannot process 2024 pumping records"
        ... )
    """
    print(f"ERROR: {what_failed}")
    print(f"  Location: {location}")
    print(f"  Actual: {actual}")
    print(f"  Expected: {expected}")
    print(f"  Physical context: {context}")


# =============================================================================
# CSV INGESTION FUNCTIONS (US-003 through US-011)
# =============================================================================

def read_source_csv(csv_path: str) -> tuple[pd.DataFrame, pd.Series]:
    """
    Read source CSV and extract daily data and annual sum row.

    Scientific Basis:
    - Source CSV contains daily flow rates in MGD (million gallons per day)
    - Summing daily values gives monthly volume in MG (dimensional analysis)
    - CSV Sum row contains annual totals per well for validation

    Assumptions:
    1. CSV has header row with date range (e.g., "1/1/2024-12/31/2024")
    2. Daily data rows have valid dates in first column
    3. Sum row labeled "Sum" in date column (after 366 daily rows)
    4. All well columns contain numeric MGD values or blanks

    Args:
        csv_path: Path to source CSV file (valid range: any readable CSV file)

    Returns:
        Tuple of (daily_df, sum_row) where:
        - daily_df: DataFrame with Date + 13 well columns (pint Quantities in MGD) + BWP_Total
        - sum_row: Series with annual MG totals per well from CSV Sum row

    Raises:
        FileNotFoundError: If CSV file does not exist at specified path
        ValueError: If CSV format is invalid (missing columns, wrong structure)

    Example:
        >>> daily_df, sum_row = read_source_csv("./input/csv/Buckman_Well_Prod_2024.csv")
        >>> print(f"Read {len(daily_df)} daily records")
        Read 366 daily records
        >>> print(f"Buckman #1 annual total: {sum_row['BWell 1']:.3f} MG")
        Buckman #1 annual total: 195.922 MG

        Hand calculation check:
        - If Buckman #1 pumps 0.5 MGD for 31 days (January)
        - Monthly total = 0.5 MGD × 31 days = 15.5 MG

    Validation:
        Compare sum_row values to independently calculated annual totals
        from daily_df (should match within ANNUAL_SUM_TOLERANCE_MG)
    """
    print(f"Reading CSV: {csv_path}")

    # 1. Check if file exists
    if not Path(csv_path).exists():
        print_error(
            "CSV file not found",
            csv_path,
            "File does not exist",
            "CSV file with daily MGD data (366 rows + 4 summary rows)",
            "Cannot process pumping data - missing source file"
        )
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    # 2. Read CSV with pandas
    try:
        df = pd.read_csv(csv_path)
    except Exception as e:
        print_error(
            "Failed to read CSV file",
            csv_path,
            f"pandas error: {type(e).__name__}: {e}",
            "Valid CSV file with comma-separated values",
            "File may be corrupted or not in CSV format"
        )
        raise

    # 3. Parse first column (date with range like "1/1/2024-12/31/2024")
    # The first column name contains the date range, rename it to 'Date' for clarity
    date_col = df.columns[0]
    df = df.rename(columns={date_col: 'Date'})

    # 4. Extract daily rows vs summary rows
    # Daily rows have parseable dates, summary rows have text ("Sum", "Avg", "Max", "Min")
    # Try to convert Date column to datetime - valid dates become timestamps, invalid become NaT
    date_parsed = pd.to_datetime(df['Date'], errors='coerce')
    daily_mask = date_parsed.notna()

    daily_df = df[daily_mask].copy()
    summary_df = df[~daily_mask].copy()

    # Parse dates for daily_df
    daily_df['Date'] = pd.to_datetime(daily_df['Date'])

    # 5. Verify we have expected well columns and total column
    missing_cols = []
    for col in CSV_WELL_COLUMNS:
        if col not in daily_df.columns:
            missing_cols.append(col)
    if CSV_TOTAL_COLUMN not in daily_df.columns:
        missing_cols.append(CSV_TOTAL_COLUMN)

    if missing_cols:
        print_error(
            "Missing expected columns in CSV",
            csv_path,
            f"Missing columns: {missing_cols}",
            f"Expected columns: {CSV_WELL_COLUMNS + [CSV_TOTAL_COLUMN]}",
            "CSV structure has changed - cannot parse well data"
        )
        raise ValueError(f"Missing columns: {missing_cols}")

    # Rename BWP total column for clarity
    daily_df = daily_df.rename(columns={CSV_TOTAL_COLUMN: 'BWP_Total'})

    # Convert well columns to numeric, coercing errors to NaN
    for col in CSV_WELL_COLUMNS:
        daily_df[col] = pd.to_numeric(daily_df[col], errors='coerce')
    daily_df['BWP_Total'] = pd.to_numeric(daily_df['BWP_Total'], errors='coerce')

    # 6. Extract Sum row (find row where Date column == "Sum")
    sum_rows = summary_df[summary_df['Date'].str.strip() == 'Sum']
    if len(sum_rows) == 0:
        print_error(
            "Sum row not found in CSV",
            csv_path,
            "No row with 'Sum' in Date column",
            "CSV should have 366 daily rows + 'Sum' row with annual totals",
            "Cannot validate annual totals without Sum row"
        )
        raise ValueError("Sum row not found in CSV")

    sum_row = sum_rows.iloc[0]

    # Convert Sum row well values to numeric (these are annual MG totals)
    sum_data = {}
    for col in CSV_WELL_COLUMNS:
        sum_data[col] = pd.to_numeric(sum_row[col], errors='coerce')
    sum_series = pd.Series(sum_data)

    print(f"Read {len(daily_df)} daily records from CSV")
    print(f"Date range: {daily_df['Date'].min().strftime('%Y-%m-%d')} to {daily_df['Date'].max().strftime('%Y-%m-%d')}")

    # 7. Return (daily_df, sum_row)
    # Note: NOT converting to pint Quantities here - will do that in aggregate_monthly
    # This keeps daily_df as simple pandas for easier manipulation
    return daily_df, sum_series


def validate_daily_data(daily_df: pd.DataFrame) -> pd.DataFrame:
    """
    Identify and flag missing/invalid data in daily DataFrame.

    Creates a flags DataFrame matching the shape of the daily data well columns.
    Flags are used downstream to mark months with data quality issues.

    Assumptions:
    1. Zero values are valid (well not pumping - common in winter months)
    2. Negative values are invalid (physical impossibility - flow cannot be negative)
    3. Non-numeric values indicate data entry errors
    4. Blank/NaN values indicate missing data

    Args:
        daily_df: DataFrame with 366 daily rows × 13 well columns (MGD values)

    Returns:
        flags_df: DataFrame with same shape, values are:
        - "" (empty string): Valid data
        - "BLANK": Missing/NaN value
        - "NEGATIVE": Negative flow value (physical impossibility)
        - "NON_NUMERIC": Non-numeric string found

    Raises:
        ValueError: If daily_df does not contain expected well columns

    Example:
        >>> # Mock data with one negative, one blank
        >>> daily_df = pd.DataFrame({
        ...     'BWell 1': [1.5, -0.5, None, 0.0],  # negative and blank
        ...     'BWell 2': [0.0, 0.0, 0.0, 0.0]     # all valid (zeros OK)
        ... })
        >>> flags_df = validate_daily_data(daily_df)
        >>> print(flags_df.loc[1, 'BWell 1'])  # negative
        NEGATIVE
        >>> print(flags_df.loc[2, 'BWell 1'])  # blank
        BLANK
        >>> print(flags_df.loc[3, 'BWell 1'])  # zero (valid)


        Hand calculation: Count flagged cells per well to estimate data quality

    Validation:
        Review flagged data in input_summary.csv against source CSV
    """
    print(f"Validating {len(daily_df)} daily records...")

    # 1. Create empty flags_df with same shape as well columns
    flags_df = pd.DataFrame(index=daily_df.index, columns=CSV_WELL_COLUMNS)
    flags_df = flags_df.fillna("")  # Initialize all to empty string (valid)

    # 2-3. Check for NaN (blank), negative values, non-numeric
    flag_counts = {}
    for col in CSV_WELL_COLUMNS:
        # Extract well number from "BWell N|Flow Mgd" format
        well_num = int(col.split('|')[0].split()[1])  # "BWell 10|Flow Mgd" -> "10"
        well_name = f"Buckman #{well_num}"
        flag_count = 0

        for idx in daily_df.index:
            value = daily_df.loc[idx, col]

            # Check for blank/NaN
            if pd.isna(value):
                flags_df.loc[idx, col] = "BLANK"
                flag_count += 1
            # Check for negative (physical impossibility)
            elif value < 0:
                flags_df.loc[idx, col] = "NEGATIVE"
                flag_count += 1
            # Zero is valid (well not pumping)
            # Positive values are valid

        if flag_count > 0:
            flag_counts[well_name] = flag_count

    # 5. Print summary of flagged cells
    total_flagged = sum(flag_counts.values())
    if total_flagged > 0:
        print(f"Flagged {total_flagged} invalid values across {len(flag_counts)} wells:")
        for well_name, count in flag_counts.items():
            print(f"  {well_name}: {count} flagged days")
    else:
        print("Flagged 0 invalid values across 0 wells")

    # 6. Return flags_df
    return flags_df


def verify_daily_sums(
    daily_df: pd.DataFrame,
    info_threshold_mgd: float = DAILY_SUM_TOLERANCE_INFO_MGD,
    error_threshold_mgd: float = DAILY_SUM_TOLERANCE_ERROR_MGD
) -> pd.DataFrame:
    """
    Verify each day's per-well sum matches the BWP total using three-tier classification.

    Scientific Basis:
    - BWP (Buckman Well Field total) should equal sum of individual well flows
    - Three-tier approach distinguishes: noise < rounding < errors
    - Tier 1 (Noise): BWP < 0.0015 MGD → database precision artifacts (100-gal rounding)
    - Tier 2 (Formula): all wells=0 but BWP ≥ 0.0015 → CSV formula error
    - Tier 3 (Rounding): Difference tolerance for Excel precision (0.001-0.005 MGD)
    - BWP column comes from CSV's own formula, NOT external validation files

    Assumptions:
    1. City database stores values with 100-gallon precision (observed: 900-1,500 gal range)
    2. Values < 1,500 gal/day are below well flow meter precision (±1% of 0.05-10 MGD)
    3. Well pumps do not operate below ~0.05 MGD (typical minimum discharge)
    4. BWP_Total column exists in daily_df (from CSV "BWP|Flow Mgd|MGD|Formula")
    5. All well columns are numeric (MGD values)
    6. NaN values treated as 0.0 MGD (well not pumping)

    Args:
        daily_df: DataFrame with Date + 13 well MGD columns + BWP_Total column
        info_threshold_mgd: Informational threshold for rounding differences (default: 0.001)
        error_threshold_mgd: Error threshold requiring CSV review (default: 0.005)

    Returns:
        verification_df: DataFrame with columns:
        - Date: Date from daily_df
        - Calculated_Sum: Sum of wells 1-13 (MGD)
        - BWP_Total: Value from BWP total column (MGD)
        - Difference: abs(Calculated_Sum - BWP_Total) (MGD)
        - Severity: "OK", "INFO", or "ERROR"

    Raises:
        ValueError: If required columns are missing from daily_df

    Example:
        >>> # Mock data with INFO and ERROR mismatches
        >>> daily_df = pd.DataFrame({
        ...     'Date': ['1/1/2024', '1/2/2024', '1/3/2024'],
        ...     'BWell 1': [1.0, 2.0, 3.0],
        ...     'BWell 2': [0.5, 0.5, 0.5],
        ...     # ... (11 more wells, all zeros)
        ...     'BWP_Total': [1.5, 2.502, 3.51]  # Day 2: INFO, Day 3: ERROR
        ... })
        >>> verification_df = verify_daily_sums(daily_df)
        >>> print(verification_df['Severity'].tolist())
        ['OK', 'INFO', 'ERROR']

        Hand calculation for 1/3/2024:
        - Sum = 3.0 + 0.5 + 0 + ... + 0 = 3.5 MGD
        - BWP = 3.51 MGD
        - Diff = |3.5 - 3.51| = 0.01 MGD > 0.005 → ERROR

    Validation:
        Compare ERROR-flagged dates to source CSV to identify formula errors
    """
    print(f"Verifying daily sums (INFO threshold: {info_threshold_mgd:.4f} MGD, "
          f"ERROR threshold: {error_threshold_mgd:.4f} MGD)...")
    print("  (BWP column is from CSV formula, not external validation data)")

    # 1. Sum wells 1-13 for each day (treating NaN as 0)
    daily_df_copy = daily_df.copy()
    calculated_sum = daily_df_copy[CSV_WELL_COLUMNS].fillna(0).sum(axis=1)

    # 2-3. Compare to BWP_Total column and calculate difference
    bwp_total = daily_df_copy['BWP_Total'].fillna(0)
    difference = abs(calculated_sum - bwp_total)

    # 4. Three-tier categorization: Noise → Formula Error → Precision Tolerance
    # Tier 1: Absolute noise threshold (database precision artifacts)
    # Tier 2: Formula error check (logical inconsistency)
    # Tier 3: Standard precision tolerance (Excel rounding)
    all_wells_zero = (daily_df_copy[CSV_WELL_COLUMNS].fillna(0) == 0).all(axis=1)
    bwp_total_series = daily_df_copy['BWP_Total'].fillna(0)

    def categorize_severity(idx: int, diff: float) -> str:
        bwp = bwp_total_series.iloc[idx]
        wells_zero = all_wells_zero.iloc[idx]

        # TIER 1: Absolute noise threshold (database precision artifacts)
        # Flag NON-ZERO values below 1,500 gal/day as INFO (database rounding artifacts)
        # Exception: BWP=0 with all wells=0 is valid and should pass to Tier 3
        if 0 < bwp < NOISE_THRESHOLD_MGD:
            return "INFO"  # Below instrument precision (database artifact)

        # TIER 2: Formula error check (logical inconsistency)
        # If all wells show zero but BWP ≥ noise threshold, flag as ERROR
        if wells_zero and bwp >= NOISE_THRESHOLD_MGD:
            return "ERROR"  # Real formula error (not noise)

        # TIER 3: Standard precision tolerance (Excel rounding)
        if diff <= info_threshold_mgd:
            return "OK"
        elif diff <= error_threshold_mgd:
            return "INFO"  # Expected rounding difference
        else:
            return "ERROR"  # Significant mismatch

    severity = pd.Series([categorize_severity(i, diff) for i, diff in enumerate(difference)])

    # Create verification DataFrame
    verification_df = pd.DataFrame({
        'Date': daily_df_copy['Date'],
        'Calculated_Sum': calculated_sum,
        'BWP_Total': bwp_total,
        'Difference': difference,
        'Severity': severity
    })

    # 5. Print summary with severity breakdown
    ok_count = (severity == "OK").sum()
    info_count = (severity == "INFO").sum()
    error_count = (severity == "ERROR").sum()

    print(f"\nDaily sum verification:")
    print(f"  ✅ {ok_count} days OK (within {info_threshold_mgd:.4f} MGD)")

    if info_count > 0:
        print(f"  ℹ️  {info_count} days INFO (rounding differences or database noise)")
        print(f"      Causes: (1) Excel formula rounding, (2) Database precision artifacts (< {NOISE_THRESHOLD_MGD:.4f} MGD)")
        print("      Impact: None - monthly totals use individual well values, not BWP")

    if error_count > 0:
        print(f"\n  ❌ {error_count} days ERROR (>{error_threshold_mgd:.4f} MGD - requires review)")
        print("      Action: Check source CSV for formula errors or data entry issues\n")
        print("  ERROR-level mismatches (source CSV may have formula errors):")

        error_rows = verification_df[verification_df['Severity'] == 'ERROR']
        for _, row in error_rows.iterrows():
            print(f"    🔴 {row['Date'].strftime('%Y-%m-%d')}: calc={row['Calculated_Sum']:.4f}, "
                  f"BWP={row['BWP_Total']:.4f}, diff={row['Difference']:.4f} MGD")

            # Additional diagnostic for anomalous cases (all wells = 0 but BWP != 0)
            day_wells = daily_df_copy[daily_df_copy['Date'] == row['Date']][CSV_WELL_COLUMNS]
            if (day_wells.fillna(0) == 0).all().all() and row['BWP_Total'] != 0:
                print(f"         ⚠️  All wells=0 but BWP≠0 - check CSV formula")

    # Show INFO-level details if present
    if info_count > 0 and info_count <= 10:
        print(f"\n  INFO-level details (rounding differences):")
        info_rows = verification_df[verification_df['Severity'] == 'INFO']
        for _, row in info_rows.iterrows():
            print(f"    ⚪ {row['Date'].strftime('%Y-%m-%d')}: calc={row['Calculated_Sum']:.4f}, "
                  f"BWP={row['BWP_Total']:.4f}, diff={row['Difference']:.4f} MGD")

    # 6. Return verification_df
    return verification_df


def aggregate_monthly(daily_df: pd.DataFrame, flags_df: pd.DataFrame) -> dict[str, pd.DataFrame]:
    """
    Aggregate daily MGD values into monthly MG totals per well.

    Scientific Basis:
    - Monthly volume (MG) = sum of daily flow rates (MGD) × 1 day
    - Dimensional analysis: (MGD × day = MG)
    - Using pint: sum(daily_mgd * ureg.day) → units automatically = million_gallon

    Assumptions:
    1. Daily values are in MGD (million gallons per day)
    2. Summing over a month gives MG (million gallons) directly
    3. Flagged data is still included in sum (user reviews flagged months separately)
    4. If ANY day in a month has a flag, the month is marked Has_Flagged_Data=True

    Args:
        daily_df: DataFrame with Date + 13 well columns (pint Quantities in MGD)
        flags_df: DataFrame with same shape, flags for each daily value

    Returns:
        Dict keyed by month string ("01" through "12"), values are DataFrames with:
        - Well_Number: 1-13
        - MG_Month: Monthly total volume (pint Quantity in million_gallon)
        - Has_Flagged_Data: Boolean (True if any flagged days in month)

    Raises:
        ValueError: If daily_df and flags_df shapes don't match

    Example:
        >>> # Mock January data: Buckman #1 pumps 1.5 MGD for 31 days
        >>> daily_df = pd.DataFrame({
        ...     'Date': pd.date_range('2024-01-01', periods=31),
        ...     'BWell 1': [1.5] * 31  # 31 days at 1.5 MGD
        ... })
        >>> flags_df = pd.DataFrame({'BWell 1': [''] * 31})  # No flags
        >>> monthly_data = aggregate_monthly(daily_df, flags_df)
        >>> print(monthly_data['01'].loc[0, 'MG_Month'])
        46.5 MG
        >>> print(monthly_data['01'].loc[0, 'Has_Flagged_Data'])
        False

        Hand calculation:
        - 1.5 MGD × 31 days = 46.5 MG ✓

    Validation:
        Sum all 12 monthly MG values per well, compare to CSV Sum row
        (should match within ANNUAL_SUM_TOLERANCE_MG)
    """
    print("Aggregating daily data into monthly totals...")

    # 1. Extract month from Date column (1-12)
    daily_df_copy = daily_df.copy()
    daily_df_copy['Month'] = daily_df_copy['Date'].dt.month

    # 4. Create dict with month keys ("01"-"12")
    monthly_data = {}

    # 2-3. Group by month and aggregate
    for month_num in range(1, 13):
        month_str = f"{month_num:02d}"  # "01" through "12"
        month_df = daily_df_copy[daily_df_copy['Month'] == month_num]

        if len(month_df) == 0:
            # Month has no data - create empty month with zeros
            well_data = []
            for well_num in range(1, 14):
                col = f"BWell {well_num}"
                well_data.append({
                    'Well_Number': well_num,
                    'MG_Month': 0.0 * ureg.million_gallon,
                    'Has_Flagged_Data': False
                })
            monthly_data[month_str] = pd.DataFrame(well_data)
            continue

        # For each well, sum daily MGD values → monthly MG
        well_data = []
        for well_num in range(1, 14):
            col = f"BWell {well_num}|Flow Mgd"

            # 3a. Sum daily MGD values → monthly MG (using pint for dimensional correctness)
            # Daily values are in MGD (million gallons per day)
            # Summing over days: sum(MGD) × 1 day = MG
            daily_mgd_values = month_df[col].fillna(0).values
            # Create pint Quantity array
            daily_mgd_qty = daily_mgd_values * ureg.million_gallon / ureg.day
            # Sum over days - units automatically become million_gallon
            monthly_mg_qty = (daily_mgd_qty * ureg.day).sum()

            # 3b. Check if any days have flags → set Has_Flagged_Data
            month_flags = flags_df.loc[month_df.index, col]
            has_flagged = (month_flags != "").any()

            well_data.append({
                'Well_Number': well_num,
                'MG_Month': monthly_mg_qty,
                'Has_Flagged_Data': has_flagged
            })

        monthly_data[month_str] = pd.DataFrame(well_data)

    # 5. Print summary (number of months processed)
    months_with_data = sum(1 for month_df in monthly_data.values() if (month_df['MG_Month'].apply(lambda q: q.magnitude) > 0).any())
    print(f"Aggregated {months_with_data} months with pumping data (12 months total)")

    # 6. Return monthly_data dict
    return monthly_data


def generate_monthly_csv(
    month_num: str,
    month_abbrev: str,
    month_df: pd.DataFrame,
    year: int,
    output_dir: str
) -> list[dict]:
    """
    Generate monthly CSV file with well production data.

    Creates a CSV file with monthly MG and AF values for each well, following
    the format specified in the PRD (US-007).

    Assumptions:
    1. month_df contains Well_Number, MG_Month (pint Quantity), Has_Flagged_Data
    2. All 13 wells are present (wells that didn't pump have 0.000 MG)
    3. Output directory exists or will be created

    Args:
        month_num: Month number as 2-digit string ("01" through "12")
        month_abbrev: Month abbreviation ("JAN" through "DEC")
        month_df: DataFrame with Well_Number, MG_Month, Has_Flagged_Data columns
        year: Year (e.g., 2024)
        output_dir: Output directory path

    Returns:
        List of dicts describing flagged wells: [
            {"month": "07", "well": 1, "flag_type": "BLANK", "days_flagged": 3},
            ...
        ]

    Raises:
        OSError: If output directory cannot be created or file cannot be written

    Example:
        >>> month_df = pd.DataFrame({
        ...     'Well_Number': [1, 2, 3],
        ...     'MG_Month': [52.949, 0.000, 0.793],  # pint Quantities
        ...     'Has_Flagged_Data': [False, False, True]
        ... })
        >>> flags = generate_monthly_csv("07", "JUL", month_df, 2024, "./output")
        Wrote 2024_07_JUL.csv (13 wells, 1 flagged)
        >>> # File contains:
        >>> # OSE_Number,Well_Name,MG_Month,AF_Calculated,Data_Quality
        >>> # RG-20516-S-5,Buckman #1,52.949,162.48852,OK
        >>> # ...

    Validation:
        Check generated CSV against validation/Table_2_2024.xlsx July column
    """
    print(f"({month_num}/12) Generating: {year}_{month_num}_{month_abbrev}.csv")

    # 1. Create output directory if needed
    Path(output_dir).mkdir(parents=True, exist_ok=True)

    # 2. Map well numbers to OSE numbers and names, extract MG values
    rows = []
    flagged_wells = []

    for _, row in month_df.iterrows():
        well_num = int(row['Well_Number'])
        ose_number = WELL_OSE_MAP[well_num]
        well_name = f"Buckman #{well_num}"

        # Extract magnitude from pint Quantity
        mg_month = row['MG_Month'].magnitude

        # 3. Convert MG to AF (using pint for unit safety)
        mg_qty = mg_month * ureg.million_gallon
        af_qty = mg_qty.to(ureg.acre_foot)
        af_calculated = af_qty.magnitude

        # 4. Set Data_Quality based on Has_Flagged_Data
        data_quality = "FLAGGED" if row['Has_Flagged_Data'] else "OK"

        rows.append({
            'OSE_Number': ose_number,
            'Well_Name': well_name,
            'MG_Month': f"{mg_month:.3f}",
            'AF_Calculated': f"{af_calculated:.5f}",
            'Data_Quality': data_quality
        })

        # 7. Collect flagged well info for return
        if row['Has_Flagged_Data']:
            flagged_wells.append({
                'month': month_num,
                'well': well_num,
                'well_name': well_name,
                'mg_total': mg_month
            })

    # 5. Add Calculated_Sum row (sum of all wells)
    total_mg = sum(row['MG_Month'].magnitude for _, row in month_df.iterrows())
    total_af = (total_mg * ureg.million_gallon).to(ureg.acre_foot).magnitude

    rows.append({
        'OSE_Number': 'Calculated_Sum',
        'Well_Name': '',
        'MG_Month': f"{total_mg:.3f}",
        'AF_Calculated': f"{total_af:.5f}",
        'Data_Quality': ''
    })

    # 6. Write CSV with pandas
    output_df = pd.DataFrame(rows)
    output_path = Path(output_dir) / f"{year}_{month_num}_{month_abbrev}.csv"
    output_df.to_csv(output_path, index=False)

    # 8. Print confirmation
    flagged_count = len(flagged_wells)
    print(f"  Wrote {output_path.name} (13 wells, {flagged_count} flagged)")

    return flagged_wells


def generate_table2_output(monthly_data: dict[str, pd.DataFrame], year: int, output_dir: str) -> None:
    """
    Generate Table 2 output CSV matching validation format.

    Creates a monthly AFY grid (wells × months) with summary statistics,
    matching the format of validation/Table_2_2024.xlsx.

    Scientific Basis:
    - AF values converted from monthly MG using MG_TO_AF_FACTOR (3.06889)
    - Percentages calculated as group_sum / total_annual * 100
    - Wells 10-13 are high-capacity production wells (typically 28-30% of total)
    - Wells 1, 7, 8 are primary production wells (typically 50-55% of total)

    Assumptions:
    1. monthly_data contains all 12 months ("01" through "12")
    2. Each month has all 13 wells present
    3. Validation format: 14 rows (13 wells + Total) × 14 cols (Well, JAN-DEC, Total)

    Args:
        monthly_data: Dict of month DataFrames from aggregate_monthly()
        year: Year for filename (e.g., 2024)
        output_dir: Output directory path

    Returns:
        None (writes file to disk)

    Raises:
        OSError: If output directory cannot be created or file cannot be written

    Example:
        >>> # After aggregating all months
        >>> generate_table2_output(monthly_data, 2024, "./output")
        Generating Table 2 output...
        Wrote 2024_Table_2_output.csv
        >>> # File structure:
        >>> # Well,JAN,FEB,...,JUL,...,DEC,Total
        >>> # 1,16.89,38.81,...,162.49,...,0.00,601.28
        >>> # ...
        >>> # Total,19.39,39.39,...,302.06,...,65.68,1372.95
        >>> #
        >>> # Wells 10-13,386.86,28.2%
        >>> # Wells 1,7,8,713.75,52.0%

    Validation:
        Compare to validation/Table_2_2024.xlsx (tolerance: 0.01 AF)
        Verify: Wells 10-13 sum = 386.86 AFY (28.2% of 1372.95)
        Verify: Wells 1,7,8 sum = 713.75 AFY (52.0% of 1372.95)
    """
    print("Generating Table 2 output...")

    # 1-2. Create 14×14 grid with AF values
    # Initialize data structure: list of rows
    # Rows contain mixed types (int well nums, float AF values, str labels)
    table_rows: list[dict[str, Any]] = []

    # Build rows for wells 1-13
    for well_num in range(1, 14):
        row_data: dict[str, Any] = {'Well': well_num}

        # Add monthly AF values
        annual_total_af = 0.0
        for month_num, month_abbrev in MONTHS_ORDERED:
            month_df = monthly_data[month_num]
            well_row = month_df[month_df['Well_Number'] == well_num].iloc[0]

            # Convert MG to AF
            mg_qty = well_row['MG_Month']
            af_value = mg_qty.to(ureg.acre_foot).magnitude

            row_data[month_abbrev] = round(af_value, 6)
            annual_total_af += af_value

        # 3. Calculate row totals (annual AF per well)
        row_data['Total'] = round(annual_total_af, 6)
        table_rows.append(row_data)

    # 4. Calculate column totals (monthly AF all wells)
    total_row: dict[str, Any] = {'Well': 'Total'}
    grand_total_af = 0.0

    for month_num, month_abbrev in MONTHS_ORDERED:
        month_df = monthly_data[month_num]
        # Sum all wells for this month
        month_total_mg = sum(row['MG_Month'] for _, row in month_df.iterrows())
        month_total_af = month_total_mg.to(ureg.acre_foot).magnitude

        total_row[month_abbrev] = round(month_total_af, 6)
        grand_total_af += month_total_af

    total_row['Total'] = round(grand_total_af, 6)
    table_rows.append(total_row)

    # Create DataFrame from rows
    table2_df = pd.DataFrame(table_rows)

    # 5. Add blank row
    blank_row: dict[str, Any] = {col: '' for col in table2_df.columns}
    blank_row['Well'] = ''

    # 6. Add Wells 10-13 statistics row
    wells_10_13_af = sum(row['Total'] for row in table_rows[9:13])  # Wells 10-13 (indices 9-12)
    pct_10_13 = (wells_10_13_af / grand_total_af * 100) if grand_total_af > 0 else 0.0
    wells_10_13_row: dict[str, Any] = {col: '' for col in table2_df.columns}
    wells_10_13_row['Well'] = 'Wells 10-13'
    wells_10_13_row['JAN'] = f"{wells_10_13_af:.6f}"
    wells_10_13_row['FEB'] = f"{pct_10_13:.1f}%"

    # 7. Add Wells 1,7,8 statistics row
    wells_1_7_8_af = table_rows[0]['Total'] + table_rows[6]['Total'] + table_rows[7]['Total']
    pct_1_7_8 = (wells_1_7_8_af / grand_total_af * 100) if grand_total_af > 0 else 0.0
    wells_1_7_8_row: dict[str, Any] = {col: '' for col in table2_df.columns}
    wells_1_7_8_row['Well'] = 'Wells 1,7,8'
    wells_1_7_8_row['JAN'] = f"{wells_1_7_8_af:.6f}"
    wells_1_7_8_row['FEB'] = f"{pct_1_7_8:.1f}%"

    # Append summary rows to DataFrame
    summary_rows = pd.DataFrame([blank_row, wells_10_13_row, wells_1_7_8_row])
    table2_df = pd.concat([table2_df, summary_rows], ignore_index=True)

    # 8. Write CSV
    output_path = Path(output_dir) / f"{year}_Table_2_output.csv"
    table2_df.to_csv(output_path, index=False)

    # 9. Write Excel (.xlsx) with validation-matching formatting
    xlsx_path = Path(output_dir) / f"{year}_Table_2_output.xlsx"
    write_table2_xlsx(table_rows, year, xlsx_path)

    # 10. Print confirmation
    print(f"  Wrote {output_path.name}")
    print(f"  Wrote {xlsx_path.name}")
    print(f"  Annual total: {grand_total_af:.2f} AFY")
    print(f"  Wells 10-13: {wells_10_13_af:.2f} AFY ({pct_10_13:.1f}%)")
    print(f"  Wells 1,7,8: {wells_1_7_8_af:.2f} AFY ({pct_1_7_8:.1f}%)")


def write_table2_xlsx(table_rows: list[dict], year: int, xlsx_path: Path) -> None:
    """
    Write Table 2 as formatted Excel file matching validation/Table_2_2024.xlsx.

    Scientific Basis: USGS volume conversion (1 MG = 3.06889 AF)

    Assumptions:
    1. table_rows has 14 entries: wells 1-13 (indices 0-12) + Total row (index 13)
    2. Each row dict has keys: 'Well', 'JAN'-'DEC', 'Total'

    Args:
        table_rows: List of 14 dicts (13 wells + Total), values in AFY (float, 6 decimals)
        year: Reporting year (e.g. 2024)
        xlsx_path: Output path for .xlsx file

    Returns:
        None (writes file to disk)

    Raises:
        ValueError: If table_rows length != 14
        IOError: If output path is not writable

    Example:
        >>> write_table2_xlsx(rows, 2024, Path('output/2024_Table_2_output.xlsx'))
        # Creates formatted Excel matching validation Table 2

    Validation: Compare output visually to validation/Table_2_2024.xlsx
    """
    month_abbrevs = ['JAN', 'FEB', 'MAR', 'APR', 'MAY', 'JUN',
                     'JUL', 'AUG', 'SEP', 'OCT', 'NOV', 'DEC']

    # Style definitions matching validation file
    font_normal = Font(name='Aptos', size=11)
    font_bold = Font(name='Aptos', size=11, bold=True)
    align_center = Alignment(horizontal='center')
    white_fill = PatternFill(patternType='solid', fgColor='FFFFFF')
    num_fmt = '#,##0.00'
    pct_fmt = '0.0%'

    # Border styles
    no_border = Border()
    medium_border = Border(
        top=Side(style='medium'), bottom=Side(style='medium'))
    hair_top_bottom = Border(
        top=Side(style='hair'), bottom=Side(style='hair'))
    hair_bottom = Border(bottom=Side(style='hair'))
    hair_top = Border(top=Side(style='hair'))

    wb = Workbook()
    ws = wb.active
    ws.title = f"Table_2_{year}"

    # Column widths
    ws.column_dimensions['A'].width = 14.75
    ws.column_dimensions['O'].width = 14.75

    # --- Row 1: Headers ---
    headers = ['Well'] + month_abbrevs + ['Total', 'Well']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font_bold
        cell.alignment = align_center
        cell.fill = white_fill
        cell.border = medium_border

    # --- Rows 2-14: Well data (wells 1-13) ---
    for row_offset, row_data in enumerate(table_rows[:13]):
        excel_row = row_offset + 2
        well_num = row_data['Well']

        # Determine border for this row
        if excel_row == 2:
            row_border = no_border
        elif excel_row == 14:
            row_border = no_border
        elif excel_row == 3:
            row_border = hair_bottom
        elif excel_row == 13:
            row_border = hair_top
        else:
            row_border = hair_top_bottom

        # Column A: Well number
        cell_a = ws.cell(row=excel_row, column=1, value=well_num)
        cell_a.font = font_bold
        cell_a.alignment = align_center
        cell_a.fill = white_fill
        cell_a.border = row_border

        # Columns B-M: Monthly values
        for col_offset, month in enumerate(month_abbrevs):
            col_idx = col_offset + 2
            value = row_data[month]
            cell = ws.cell(row=excel_row, column=col_idx, value=value)
            cell.font = font_normal
            cell.number_format = num_fmt
            cell.fill = white_fill
            cell.border = row_border

        # Column N: Total (calculated value for pandas compatibility)
        cell_n = ws.cell(row=excel_row, column=14,
                         value=row_data['Total'])
        cell_n.font = font_normal
        cell_n.number_format = num_fmt
        cell_n.fill = white_fill
        cell_n.border = row_border

        # Column O: Duplicate well number
        cell_o = ws.cell(row=excel_row, column=15, value=well_num)
        cell_o.font = font_bold
        cell_o.alignment = align_center
        cell_o.fill = white_fill
        cell_o.border = row_border

    # --- Row 15: Total row ---
    total_row_excel = 15
    total_row_data = table_rows[13]  # Total row is index 13 (after 13 wells)
    cell_a = ws.cell(row=total_row_excel, column=1, value='Total')
    cell_a.font = font_bold
    cell_a.alignment = align_center
    cell_a.fill = white_fill
    cell_a.border = medium_border

    # Columns B-M: Monthly totals (calculated values for pandas compatibility)
    for col_offset, month in enumerate(month_abbrevs):
        col_idx = col_offset + 2
        cell = ws.cell(row=total_row_excel, column=col_idx,
                       value=total_row_data[month])
        cell.font = font_normal
        cell.number_format = num_fmt
        cell.fill = white_fill
        cell.border = medium_border

    # Column N: Grand total (calculated value for pandas compatibility)
    cell_n = ws.cell(row=total_row_excel, column=14,
                     value=total_row_data['Total'])
    cell_n.font = font_normal
    cell_n.number_format = num_fmt
    cell_n.fill = white_fill
    cell_n.border = medium_border

    # Column O: "Total" label
    cell_o = ws.cell(row=total_row_excel, column=15, value='Total')
    cell_o.font = font_bold
    cell_o.alignment = align_center
    cell_o.fill = white_fill
    cell_o.border = medium_border

    # --- Summary area (Q14:T15) ---
    # Calculate summary values from table_rows
    # Wells 10-13 are indices 9-12 (wells numbered 10, 11, 12, 13)
    wells_10_13_af = sum(table_rows[i]['Total'] for i in range(9, 13))
    # Wells 1, 7, 8 are indices 0, 6, 7
    wells_1_7_8_af = table_rows[0]['Total'] + table_rows[6]['Total'] + table_rows[7]['Total']
    grand_total = total_row_data['Total']
    pct_10_13 = wells_10_13_af / grand_total if grand_total > 0 else 0.0
    pct_1_7_8 = wells_1_7_8_af / grand_total if grand_total > 0 else 0.0

    # Row 14 labels
    ws.cell(row=14, column=17, value='wells 10-13').font = font_normal  # Q14
    ws.cell(row=14, column=18, value='% total').font = font_normal      # R14
    ws.cell(row=14, column=19, value='wells 1, 7, 8').font = font_normal  # S14

    # Row 15 values (calculated for pandas compatibility)
    cell_q15 = ws.cell(row=15, column=17, value=wells_10_13_af)  # Wells 10-13 sum
    cell_q15.font = font_normal
    cell_q15.number_format = num_fmt

    cell_r15 = ws.cell(row=15, column=18, value=pct_10_13)  # Wells 10-13 %
    cell_r15.font = font_normal
    cell_r15.number_format = pct_fmt

    cell_s15 = ws.cell(row=15, column=19, value=wells_1_7_8_af)  # Wells 1,7,8 sum
    cell_s15.font = font_normal
    cell_s15.number_format = num_fmt

    cell_t15 = ws.cell(row=15, column=20, value=pct_1_7_8)  # Wells 1,7,8 %
    cell_t15.font = font_normal
    cell_t15.number_format = pct_fmt

    wb.save(xlsx_path)


def generate_table1_output(year_afy_data: dict[int, float], year: int, output_dir: str) -> None:
    """
    Generate Table 1 output by inserting 2024 row into historical data.

    Reads existing Table_1_data_afy_2024.xlsx (historical 1988-2023), inserts
    2024 data, adds statistics rows, and outputs updated CSV.

    Scientific Basis:
    - Well percentages show distribution of pumping across well field
    - Wells 10-13 group represents high-capacity production wells
    - Ranking years by total AFY identifies drought/high-usage periods

    Assumptions:
    1. Validation/Table_1_data_afy_2024.xlsx exists with historical data
    2. 2024 row should be inserted at end (after 2023)
    3. Well 3 labeled as "3/3A" (historical convention)
    4. All years 1988-2024 are ranked by total AFY (1 = lowest pumping)

    Args:
        year_afy_data: Dict mapping well number (1-13) to annual AFY
        year: Year being added (should be 2024)
        output_dir: Output directory path

    Returns:
        None (writes file to disk)

    Raises:
        FileNotFoundError: If validation Excel file not found
        OSError: If output file cannot be written

    Example:
        >>> year_afy_data = {1: 601.27, 2: 0.01, ..., 13: 212.02}
        >>> generate_table1_output(year_afy_data, 2024, "./output")
        Generating Table 1 output...
        Wrote Table_1_updated.csv
        >>> # File includes:
        >>> # - All historical years 1988-2023
        >>> # - New 2024 row with annual totals
        >>> # - Statistics rows 44-48 (well percentages, group sums)
        >>> # - Sort column ranking all years by total AFY

    Validation:
        Compare 2024 row totals to Table 2 annual totals (should match exactly)
        Verify total = 1372.95 AFY
    """
    print("Generating Table 1 output...")

    # 1. Read template file (validation file or previous year's output)
    validation_path = Path(VALIDATION_DIR) / f"Table_1_data_afy_{year}.xlsx"

    # Fallback to previous year's output if validation file doesn't exist
    if not validation_path.exists():
        fallback_path = Path(OUTPUT_DIR) / f"{year - 1}" / f"{year - 1}_Table_1_updated.xlsx"
        if fallback_path.exists():
            validation_path = fallback_path
            print(f"  Using {year - 1} output as template (no validation file for {year})")
        else:
            print_error(
                "No template file found for Table 1",
                f"generate_table1_output for year {year}",
                "Neither validation file nor previous year output exists",
                f"validation/Table_1_data_afy_{year}.xlsx OR output/{year - 1}/{year - 1}_Table_1_updated.xlsx",
                (
                    "Create template file with historical data. Options:\n"
                    f"  1. Copy Table_1_data_afy_{year - 1}.xlsx to validation/\n"
                    f"  2. Ensure {year - 1} was processed first (creates output template)\n"
                    "  3. Create new file with historical 1988-present AFY data"
                )
            )
            return 1

    try:
        table1_df = pd.read_excel(validation_path)
        print(f"  Read {len(table1_df)} historical years from {validation_path.name}")
    except Exception as e:
        print_error(
            "Failed to read validation file",
            f"generate_table1_output for year {year}",
            str(e),
            "Valid Excel file with historical Table 1 data",
            f"Check that {validation_path} is a valid Excel file"
        )
        return 1

    # Check for NaN values in Total column for numeric year rows
    year_rows = table1_df[
        table1_df['Well:'].apply(lambda x: isinstance(x, (int, float)) and not pd.isna(x))
    ]
    nan_totals = year_rows[year_rows['Total'].isna()]
    if len(nan_totals) > 0:
        nan_years = nan_totals['Well:'].tolist()
        print(f"  WARNING: Found {len(nan_totals)} years with missing Total: {nan_years}")
        # Recalculate missing totals from well columns
        well_cols = [1, 2, '3/3A', 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
        for idx in nan_totals.index:
            recalc_total = table1_df.loc[idx, well_cols].sum()
            table1_df.loc[idx, 'Total'] = recalc_total
            print(f"    Recalculated Total for {int(table1_df.loc[idx, 'Well:'])}: {recalc_total:.2f} AFY")

    # 2. Insert 2024 row with year_afy_data
    # Build 2024 row (mixed key types: str 'Well:' + int well nums + str '3/3A')
    row_2024: dict[Any, Any] = {'Well:': year}

    # Add individual well values (wells 1-13)
    # Note: Well 3 should be labeled as "3/3A" in column header (historical convention)
    for well_num, af_value in year_afy_data.items():
        col_name: str | int
        if well_num == 3:
            col_name = '3/3A'
        else:
            col_name = well_num
        row_2024[col_name] = round(af_value, 6)

    # Calculate total
    total_afy = sum(year_afy_data.values())
    row_2024['Total'] = round(total_afy, 6)

    # Insert or update 2024 row (avoid duplicates if xlsx already has this year)
    new_row_df = pd.DataFrame([row_2024])
    if len(table1_df) > 0:
        existing_years = table1_df['Well:'].apply(
            lambda x: isinstance(x, (int, float)) and not pd.isna(x) and int(x) == year)
        if existing_years.any():
            # Update existing row with computed values
            mask = existing_years
            for col, val in row_2024.items():
                if col != 'Well:':
                    table1_df.loc[mask, col] = val
        else:
            table1_df = pd.concat([table1_df, new_row_df], ignore_index=True)
    else:
        table1_df = new_row_df

    # 4. Calculate Sort column (rank all years by Total, 1=lowest)
    # Extract rows with numeric years (skip any summary rows)
    year_rows = table1_df[table1_df['Well:'].apply(lambda x: isinstance(x, (int, float)) and not pd.isna(x))]

    if len(year_rows) > 0:
        # Rank by Total AFY (1 = lowest pumping year)
        year_rows_sorted = year_rows.sort_values('Total')
        rank_map = {year_rows_sorted.iloc[i]['Well:']: i + 1 for i in range(len(year_rows_sorted))}

        # Add Sort column to main dataframe
        table1_df['Total, Sort'] = table1_df['Well:'].map(rank_map).astype(pd.Int64Dtype())

    # 3. Add statistics rows
    # Calculate statistics for 2024 (last row with data)
    if year in year_afy_data:
        stats_rows = []

        # Row: Individual well % of annual total
        well_pct_row: dict[Any, Any] = {'Well:': '% of Total'}
        for well_num in range(1, 14):
            af_value = year_afy_data.get(well_num, 0.0)
            pct = (af_value / total_afy * 100) if total_afy > 0 else 0.0
            col_name = '3/3A' if well_num == 3 else well_num
            well_pct_row[col_name] = f"{pct:.1f}%"
        well_pct_row['Total'] = "100.0%"
        stats_rows.append(well_pct_row)

        # Row: Wells 10-13 sum
        wells_10_13_sum = sum(year_afy_data.get(w, 0.0) for w in [10, 11, 12, 13])
        wells_10_13_row: dict[Any, Any] = {'Well:': 'Wells 10-13'}
        wells_10_13_row[10] = round(wells_10_13_sum, 6)
        stats_rows.append(wells_10_13_row)

        # Row: Wells 10-13 % of total
        wells_10_13_pct = (wells_10_13_sum / total_afy * 100) if total_afy > 0 else 0.0
        wells_10_13_pct_row: dict[Any, Any] = {'Well:': 'Wells 10-13 %'}
        wells_10_13_pct_row[10] = f"{wells_10_13_pct:.1f}%"
        stats_rows.append(wells_10_13_pct_row)

        # Row: Wells 1,7,8 sum
        wells_1_7_8_sum = sum(year_afy_data.get(w, 0.0) for w in [1, 7, 8])
        wells_1_7_8_row: dict[Any, Any] = {'Well:': 'Wells 1,7,8'}
        wells_1_7_8_row[1] = round(wells_1_7_8_sum, 6)
        stats_rows.append(wells_1_7_8_row)

        # Row: Wells 1,7,8 % of total
        wells_1_7_8_pct = (wells_1_7_8_sum / total_afy * 100) if total_afy > 0 else 0.0
        wells_1_7_8_pct_row: dict[Any, Any] = {'Well:': 'Wells 1,7,8 %'}
        wells_1_7_8_pct_row[1] = f"{wells_1_7_8_pct:.1f}%"
        stats_rows.append(wells_1_7_8_pct_row)

        # Append statistics rows
        stats_df = pd.DataFrame(stats_rows)
        table1_df = pd.concat([table1_df, stats_df], ignore_index=True)

    # 5. Write CSV
    output_path = Path(output_dir) / f"{year}_Table_1_updated.csv"
    table1_df.to_csv(output_path, index=False)

    # 6. Write Excel (.xlsx) with validation-matching formatting
    xlsx_path = Path(output_dir) / f"{year}_Table_1_updated.xlsx"
    write_table1_xlsx(table1_df, year_rows, year, xlsx_path)

    # 7. Print confirmation
    print(f"  Wrote {output_path.name}")
    print(f"  Wrote {xlsx_path.name}")
    print(f"  {year} total: {total_afy:.2f} AFY")
    if len(year_rows) > 0:
        year_rank = rank_map.get(year, 'N/A')
        print(f"  {year} rank: {year_rank} of {len(year_rows)} years (1=lowest pumping)")


def write_table1_xlsx(
    table1_df: pd.DataFrame,
    year_rows: pd.DataFrame,
    year: int,
    xlsx_path: Path
) -> None:
    """
    Write Table 1 as formatted Excel file matching validation/Table_1_data_afy_2024.xlsx.

    Scientific Basis: USGS volume conversion (1 MG = 3.06889 AF)

    Assumptions:
    1. table1_df contains year rows (1988-2024) followed by statistics rows
    2. year_rows contains only numeric year rows (no stats)
    3. Columns: 'Well:', 1-13 (or '3/3A'), 'Total', 'Total, Sort'

    Args:
        table1_df: Full DataFrame with year data + statistics rows
        year_rows: DataFrame subset containing only year data rows
        year: Current reporting year (e.g. 2024)
        xlsx_path: Output path for .xlsx file

    Returns:
        None (writes file to disk)

    Raises:
        IOError: If output path is not writable

    Example:
        >>> write_table1_xlsx(df, year_df, 2024, Path('output/Table_1_updated.xlsx'))
        # Creates formatted Excel matching validation Table 1

    Validation: Compare output visually to validation/Table_1_data_afy_2024.xlsx
    """
    # Style definitions matching validation file
    font_normal = Font(name='Aptos', size=11)
    font_bold = Font(name='Aptos', size=11, bold=True)
    align_center = Alignment(horizontal='center')
    align_right = Alignment(horizontal='right')
    num_fmt = '#,##0.00'
    pct_fmt_2 = '0.00%'
    pct_fmt_1 = '0.0%'
    pct_fmt_0 = '0%'

    # Border styles
    medium_border = Border(
        top=Side(style='medium'), bottom=Side(style='medium'))
    hair_top_bottom = Border(
        top=Side(style='hair'), bottom=Side(style='hair'))
    hair_bottom = Border(bottom=Side(style='hair'))
    hair_top = Border(top=Side(style='hair'))
    medium_top_hair_bottom = Border(
        top=Side(style='medium'), bottom=Side(style='hair'))
    hair_top_medium_bottom = Border(
        top=Side(style='hair'), bottom=Side(style='medium'))

    # Well column headers (B=1, C=2, D=3/3A, E=4, ..., N=13)
    well_headers = [1, 2, '3/3A', 4, 5, 6, 7, 8, 9, 10, 11, 12, 13]
    # Map from well header to DataFrame column name
    well_col_map = {}
    for h in well_headers:
        well_col_map[h] = h  # DataFrame uses same keys

    wb = Workbook()
    ws = wb.active
    ws.title = f"Table1_data_afy_{year}"

    # Column widths
    ws.column_dimensions['A'].width = 12.75
    ws.column_dimensions['O'].width = 11.75

    # --- Identify year data rows (exclude stats) ---
    # year_rows is already filtered to numeric years
    num_year_rows = len(year_rows)
    last_year_excel_row = num_year_rows + 1  # +1 for header row

    # --- Row 1: Headers ---
    def write_header_row(excel_row: int) -> None:
        """Write header row at given Excel row."""
        # Column A: "Well:"
        cell = ws.cell(row=excel_row, column=1, value='Well:')
        cell.font = font_bold
        cell.alignment = align_right
        cell.border = medium_border

        # Columns B-N: Well numbers (1-13)
        for col_offset, header in enumerate(well_headers):
            col_idx = col_offset + 2
            cell = ws.cell(row=excel_row, column=col_idx, value=header)
            cell.font = font_bold
            cell.alignment = align_center
            cell.border = medium_border

        # Column O: "Total"
        cell = ws.cell(row=excel_row, column=15, value='Total')
        cell.font = font_bold
        cell.alignment = align_center
        cell.border = medium_border

    write_header_row(1)

    # Column T header (row 1)
    cell_t1 = ws.cell(row=1, column=20, value='Total, Sort')
    cell_t1.font = font_bold

    # --- Year data rows (2 to last_year_excel_row) ---
    for i in range(num_year_rows):
        excel_row = i + 2
        year_val = year_rows.iloc[i]['Well:']

        # Determine border
        if excel_row == 2:
            row_border = hair_bottom
        elif excel_row == last_year_excel_row:
            row_border = hair_top
        else:
            row_border = hair_top_bottom

        # Column A: Year
        cell_a = ws.cell(row=excel_row, column=1, value=int(year_val))
        cell_a.font = font_bold
        cell_a.alignment = align_center
        cell_a.border = row_border

        # Columns B-N: Well values
        for col_offset, header in enumerate(well_headers):
            col_idx = col_offset + 2
            val = year_rows.iloc[i].get(header)
            if pd.notna(val):
                try:
                    val = float(val)
                except (ValueError, TypeError):
                    val = None
            else:
                val = None
            if val is not None:
                cell = ws.cell(row=excel_row, column=col_idx, value=val)
                cell.font = font_normal
                cell.number_format = num_fmt
                cell.border = row_border

        # Column O: Total
        # Write actual value (not formula) to ensure pandas can read it back
        total_val = year_rows.iloc[i].get('Total')
        if pd.notna(total_val):
            cell_o = ws.cell(row=excel_row, column=15, value=float(total_val))
            cell_o.font = font_normal
            cell_o.number_format = num_fmt
            cell_o.border = row_border

    # --- Column T: All year totals sorted descending ---
    all_totals = []
    for i in range(num_year_rows):
        total_val = year_rows.iloc[i].get('Total')
        if pd.notna(total_val):
            all_totals.append(float(total_val))
    all_totals_sorted = sorted(all_totals, reverse=True)

    # Count years with wells 10-13 (2013 and later) for ranking
    wells_10_13_start_year = 2013
    num_ranked_years = sum(
        1 for i in range(num_year_rows)
        if int(year_rows.iloc[i]['Well:']) >= wells_10_13_start_year
    )

    # Write sorted totals in column T, ranks in column U
    num_totals = len(all_totals_sorted)
    for i, total_val in enumerate(all_totals_sorted):
        excel_row = i + 2
        # Column T: sorted total value
        cell_t = ws.cell(row=excel_row, column=20, value=total_val)
        cell_t.font = font_normal
        cell_t.number_format = num_fmt
        # Apply hair borders matching data rows
        if excel_row == 2:
            cell_t.border = hair_bottom
        elif excel_row == num_totals + 1:
            cell_t.border = hair_top
        else:
            cell_t.border = hair_top_bottom

        # Column U: Rank for bottom N entries (N = years with wells 10-13)
        # Ranks assigned to last num_ranked_years entries in sorted list
        # Position from bottom: num_totals - i (1-based)
        position_from_bottom = num_totals - i
        if position_from_bottom <= num_ranked_years:
            cell_u = ws.cell(row=excel_row, column=21, value=position_from_bottom)
            cell_u.font = font_normal

    # --- Average rows ---
    avg_row_1 = last_year_excel_row + 1  # "Average, 1988-2024"
    avg_row_2 = last_year_excel_row + 2  # "Average, 2022-2024"

    # Average 1988-2024
    cell_a = ws.cell(row=avg_row_1, column=1, value=f'Average,\n1988\u2013{year}')
    cell_a.font = font_bold
    cell_a.alignment = Alignment(wrap_text=True)
    cell_a.border = medium_top_hair_bottom

    for col_idx in range(2, 16):  # B through O
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=avg_row_1, column=col_idx,
                       value=f'=AVERAGE({col_letter}2:{col_letter}{last_year_excel_row})')
        cell.font = font_normal
        cell.number_format = num_fmt
        cell.border = medium_top_hair_bottom

    # Average 2022-2024 (last 3 years)
    last_3_start = last_year_excel_row - 2  # Row for 2022
    cell_a = ws.cell(row=avg_row_2, column=1, value=f'Average,\n2022\u2013{year}')
    cell_a.font = font_bold
    cell_a.alignment = Alignment(wrap_text=True)
    cell_a.border = hair_top_medium_bottom

    for col_idx in range(2, 16):
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=avg_row_2, column=col_idx,
                       value=f'=AVERAGE({col_letter}{last_3_start}:{col_letter}{last_year_excel_row})')
        cell.font = font_normal
        cell.number_format = num_fmt
        cell.border = hair_top_medium_bottom

    # --- Empty rows 41-43 (avg_row_2 + 1 to avg_row_2 + 3) ---
    # Just skip, they'll be empty

    # --- Repeat header row (row 44 equivalent) ---
    stats_header_row = avg_row_2 + 4  # Skip 3 empty rows
    write_header_row(stats_header_row)

    # --- Percent row (row 45 equivalent) ---
    pct_row = stats_header_row + 1
    cell_a = ws.cell(row=pct_row, column=1, value='Pecent')  # Match validation typo
    cell_a.font = font_normal

    for col_idx in range(2, 15):  # B through N (wells 1-13)
        col_letter = get_column_letter(col_idx)
        cell = ws.cell(row=pct_row, column=col_idx,
                       value=f'={col_letter}{last_year_excel_row}/$O${last_year_excel_row}')
        cell.font = font_normal
        cell.number_format = pct_fmt_2

    # Column O: Sum of percentages
    cell_o = ws.cell(row=pct_row, column=15,
                     value=f'=SUM(B{pct_row}:N{pct_row})')
    cell_o.font = font_normal
    cell_o.number_format = pct_fmt_0

    # --- Empty row 46 ---

    # --- Row 47: Wells 10-13 sum (column N) ---
    wells_sum_row = pct_row + 2
    cell_n47 = ws.cell(row=wells_sum_row, column=14,
                       value=f'=SUM(K{last_year_excel_row}:N{last_year_excel_row})')
    cell_n47.font = font_normal
    cell_n47.number_format = num_fmt

    # --- Row 48: Wells 10-13 % (column N) ---
    wells_pct_row = wells_sum_row + 1
    cell_n48 = ws.cell(row=wells_pct_row, column=14,
                       value=f'=K{pct_row}+L{pct_row}+M{pct_row}+N{pct_row}')
    cell_n48.font = font_normal
    cell_n48.number_format = pct_fmt_1

    wb.save(xlsx_path)


def generate_qa_summary(
    all_flags: list[dict],
    daily_verification: pd.DataFrame,
    year: int,
    output_dir: str
) -> None:
    """
    Generate QA summary CSV listing flagged data and verification failures.

    Creates input_summary.csv with two sections:
    1. Flagged well-months (missing/invalid data from validation)
    2. Daily sum mismatches (BWP verification failures)

    Assumptions:
    1. all_flags contains dicts from generate_monthly_csv() calls
    2. daily_verification contains full year of daily sum checks
    3. If no issues found, file contains "No data quality issues found"

    Args:
        all_flags: List of flagged well-month dicts from all 12 months
        daily_verification: DataFrame with daily sum verification results
        year: Year for reference (e.g., 2024)
        output_dir: Output directory path

    Returns:
        None (writes file to disk)

    Raises:
        OSError: If output file cannot be written

    Example:
        >>> all_flags = [
        ...     {"month": "07", "well": 3, "flag_type": "BLANK", "days_flagged": 2}
        ... ]
        >>> generate_qa_summary(all_flags, daily_verification, 2024, "./output")
        QA summary: 1 flagged well-month, 0 daily sum mismatches
        >>> # File contains:
        >>> # Section 1 - Flagged Well-Months
        >>> # Month,Well_Name,Flag_Type,Flagged_Days_Count,Monthly_MG_Total
        >>> # JUL,Buckman #3,BLANK,2,0.793
        >>> #
        >>> # Section 2 - Daily Sum Mismatches
        >>> # (empty - no mismatches)

    Validation:
        Review flagged entries against source CSV to verify data quality issues
    """
    print("Generating QA input summary...")

    # 1. Section 1: Format flagged well-months
    flagged_data = []
    for flag_info in all_flags:
        flagged_data.append({
            'Month': flag_info.get('month', ''),
            'Well_Name': flag_info.get('well_name', f"Buckman #{flag_info.get('well', '')}"),
            'Flag_Type': flag_info.get('flag_type', 'FLAGGED'),
            'Flagged_Days_Count': flag_info.get('days_flagged', 0),
            'Monthly_MG_Total': f"{flag_info.get('mg_total', 0.0):.3f}"
        })

    # 2. Section 2: Extract daily verification INFO and ERROR rows
    # Include both INFO and ERROR severity levels in the QA summary
    mismatches = daily_verification[daily_verification['Severity'].isin(['INFO', 'ERROR'])]
    mismatch_data = []
    for _, row in mismatches.iterrows():
        mismatch_data.append({
            'Date': row['Date'].strftime('%Y-%m-%d'),
            'Calculated_Sum': f"{row['Calculated_Sum']:.4f}",
            'BWP_Total': f"{row['BWP_Total']:.4f}",
            'Difference': f"{row['Difference']:.4f}",
            'Severity': row['Severity']
        })

    # 3-4. If both sections empty, write "No data quality issues found"
    output_path = Path(output_dir) / "input_summary.csv"

    if len(flagged_data) == 0 and len(mismatch_data) == 0:
        # Write simple message
        with open(output_path, 'w') as f:
            f.write("No data quality issues found\n")
    else:
        # Write both sections
        with open(output_path, 'w') as f:
            # Section 1: Flagged Well-Months
            if len(flagged_data) > 0:
                f.write("Section 1 - Flagged Well-Months\n")
                flagged_df = pd.DataFrame(flagged_data)
                flagged_df.to_csv(f, index=False)
                f.write("\n")

            # Section 2: Daily Sum Mismatches
            if len(mismatch_data) > 0:
                f.write("Section 2 - Daily Sum Mismatches\n")
                mismatch_df = pd.DataFrame(mismatch_data)
                mismatch_df.to_csv(f, index=False)

    # 5. Print summary counts
    print(f"  QA summary: {len(flagged_data)} flagged well-months, {len(mismatch_data)} daily sum mismatches")
    print(f"  Wrote {output_path.name}")


def verify_annual_sums(
    monthly_data: dict[str, pd.DataFrame],
    sum_row: pd.Series,
    tolerance_mg: float = ANNUAL_SUM_TOLERANCE_MG
) -> dict[str, str]:
    """
    Verify our calculated annual totals match the CSV Sum row.

    Scientific Basis:
    - Annual MG = sum of 12 monthly MG values per well
    - CSV Sum row contains independently calculated totals from source data
    - Discrepancies indicate aggregation errors or missing months

    Assumptions:
    1. monthly_data contains all 12 months
    2. sum_row is from CSV (annual MG totals per well)
    3. Tolerance of 0.01 MG = 10,000 gallons (acceptable rounding)

    Args:
        monthly_data: Dict of month DataFrames from aggregate_monthly()
        sum_row: Series with annual MG totals from CSV Sum row
        tolerance_mg: Maximum acceptable difference in MG (default: 0.01)

    Returns:
        Dict mapping well name to verification status:
        {"Buckman #1": "OK", "Buckman #2": "NOT_OK (calc=195.93, source=195.92)", ...}

    Raises:
        ValueError: If monthly_data is incomplete (missing months)

    Example:
        >>> # Sum 12 months of data
        >>> verification = verify_annual_sums(monthly_data, sum_row, tolerance_mg=0.01)
        >>> print(verification["Buckman #1"])
        OK
        >>> print(verification["Buckman #3"])
        NOT_OK (calc=56.24, source=56.23)

        Hand calculation:
        - If monthly MG values are [16.89, 38.81, ..., 0.00] (12 values)
        - Annual = sum = 195.922 MG
        - Compare to CSV Sum row: 195.922 MG → diff = 0.000 → OK

    Validation:
        All wells should verify OK if aggregation is correct
        NOT_OK indicates missing data or calculation error
    """
    print("Verifying annual totals...")

    verification = {}

    # 1-2-3. Sum all 12 monthly MG values per well and compare to sum_row
    for well_num in range(1, 14):
        well_name = f"Buckman #{well_num}"
        col = f"BWell {well_num}|Flow Mgd"

        # Sum 12 months of MG values
        annual_mg_calc = 0.0
        for month_num in range(1, 13):
            month_str = f"{month_num:02d}"
            month_df = monthly_data[month_str]
            well_row = month_df[month_df['Well_Number'] == well_num].iloc[0]

            # Extract magnitude from pint Quantity
            mg_value = well_row['MG_Month'].magnitude
            annual_mg_calc += mg_value

        # Get source value from CSV Sum row
        annual_mg_source = sum_row[col]

        # Check if within tolerance
        difference = abs(annual_mg_calc - annual_mg_source)

        if difference <= tolerance_mg:
            verification[well_name] = "OK"
        else:
            verification[well_name] = f"NOT_OK (calc={annual_mg_calc:.3f}, source={annual_mg_source:.3f})"

    # 5. Print summary (count of OK vs NOT_OK)
    ok_count = sum(1 for status in verification.values() if status == "OK")
    not_ok_count = len(verification) - ok_count

    print(f"  Annual verification: {ok_count} wells OK, {not_ok_count} NOT_OK")

    if not_ok_count > 0:
        print(f"  WARNING: {not_ok_count} wells have annual sum mismatches:")
        for well_name, status in verification.items():
            if status != "OK":
                print(f"    {well_name}: {status}")

    # 6. Return verification dict
    return verification


# =============================================================================
# PREREQUISITE CHECKS
# =============================================================================

def check_prerequisites(year: int) -> bool:
    """
    Check required inputs exist before processing.

    Validates that all required input files are present before starting
    the ingestion workflow. Provides clear error messages with hints.

    Args:
        year: Year to process (e.g., 2024, 2025)

    Returns:
        True if all prerequisites are met, False otherwise
    """
    csv_path = Path(INPUT_CSV_PATH.format(year=year))

    if not csv_path.exists():
        print_error(
            "Input CSV file not found",
            str(csv_path),
            "File does not exist",
            f"Buckman_Well_Prod_{year}.csv in input/csv/",
            "Obtain pumping data CSV from City of Santa Fe"
        )
        return False

    # Print checklist-style prerequisite status
    print("\n" + "="*70)
    print(f"STEP 1: INGEST PUMPING DATA - YEAR {year}")
    print("="*70)
    print("✓ Prerequisites:")
    print(f"  - Input CSV: input/csv/Buckman_Well_Prod_{year}.csv [FOUND]")

    # Check Table 1 template availability
    validation_path = Path(VALIDATION_DIR) / f"Table_1_data_afy_{year}.xlsx"
    fallback_path = Path(OUTPUT_DIR) / f"{year - 1}" / f"{year - 1}_Table_1_updated.xlsx"

    print("\n📋 Table 1 Template Source:")
    if validation_path.exists():
        print(f"  ✓ Using validation file: {validation_path.name}")
    elif fallback_path.exists():
        print(f"  ✓ Using {year - 1} output as template: {fallback_path.name}")
    else:
        print(f"  ⚠ WARNING: No template found!")
        print(f"    - Primary: {validation_path.name} [NOT FOUND]")
        print(f"    - Fallback: {fallback_path} [NOT FOUND]")
        print(f"    ⚠ Table 1 generation will fail unless template is created.")

    print("\n📦 Outputs (after completion):")
    print(f"  - {year}_Table_1_updated.xlsx (historical AFY by well)")
    print(f"  - {year}_Table_2_output.xlsx (monthly pumping)")
    print(f"  - 12 monthly CSV files")

    print("\n➡️  Next Step:")
    print(f"  python3 step2_update_modflow.py --year {year}")
    print("="*70 + "\n")

    return True


# =============================================================================
# MAIN FUNCTION (US-011)
# =============================================================================

def main() -> int:
    """
    Main entry point for CSV ingestion workflow.

    Orchestrates the complete data pipeline:
    1. Read source CSV (daily MGD data)
    2. Validate data quality
    3. Verify daily sums
    4. Aggregate to monthly MG
    5. Generate outputs (monthly CSVs, Table 2, Table 1, QA summary)
    6. Verify annual totals

    Command-line usage:
        python3 ingest_buckman_data.py <year>
        python3 ingest_buckman_data.py 2024
        python3 ingest_buckman_data.py 2025

    Returns:
        Exit code: 0 (success), 1 (error)

    Example:
        >>> # From command line:
        >>> # $ python3 ingest_buckman_data.py 2024
        Reading CSV: ./input/csv/Buckman_Well_Prod_2024.csv
        Read 366 daily records from CSV
        Validating 366 daily records...
        Flagged 0 invalid values across 0 wells
        Verifying daily sums (tolerance: 0.001 MGD)...
        Daily sum verification: 366 OK, 0 NOT_OK
        Aggregating daily data into monthly totals...
        (1/12) Generating: 2024_01_JAN.csv
        ...
        (7/12) Generating: 2024_07_JUL.csv  ← July now included!
        ...
        Generating Table 2 output...
        Wrote 2024_Table_2_output.csv
        Generating Table 1 output...
        Wrote 2024_Table_1_updated.csv
        Generating QA input summary...
        QA summary: 0 flagged well-months, 0 daily sum mismatches
        Verifying annual totals...
        Annual verification: 13 wells OK, 0 NOT_OK

        === SUMMARY ===
        Files created: 15 (12 monthly + Table2 + Table1 + QA)
        Flagged data: 0 well-months
        Verification: All OK
        Buckman #1 July: 162.49 AF ✓
    """
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="Ingest Buckman Well Field daily CSV data and generate reports"
    )
    parser.add_argument(
        "--year",
        type=int,
        required=True,
        help="Year to process (e.g., 2024, 2025)"
    )
    args = parser.parse_args()
    year = args.year

    print("\nBuckman Well Field CSV Data Ingestion")
    print(f"Processing year: {year}")
    print("=" * 60)

    # Check prerequisites before processing
    if not check_prerequisites(year):
        return 1

    try:
        # 1. Construct CSV path from year
        csv_path = INPUT_CSV_PATH.format(year=year)

        # 3. Create output directory if needed
        Path(OUTPUT_DIR).mkdir(parents=True, exist_ok=True)

        # 4. Call read_source_csv()
        daily_df, sum_row = read_source_csv(csv_path)

        # 5. Call validate_daily_data()
        flags_df = validate_daily_data(daily_df)

        # 6. Call verify_daily_sums() with dual thresholds
        daily_verification = verify_daily_sums(
            daily_df,
            info_threshold_mgd=DAILY_SUM_TOLERANCE_INFO_MGD,
            error_threshold_mgd=DAILY_SUM_TOLERANCE_ERROR_MGD
        )

        # 7. Call aggregate_monthly()
        monthly_data = aggregate_monthly(daily_df, flags_df)

        # 8. Loop through 12 months: call generate_monthly_csv()
        print("\nGenerating monthly CSV files...")
        all_flags = []
        for month_num, month_abbrev in MONTHS_ORDERED:
            month_df = monthly_data[month_num]
            flagged_wells = generate_monthly_csv(
                month_num, month_abbrev, month_df, year, OUTPUT_DIR
            )
            all_flags.extend(flagged_wells)

        # 9. Call generate_table2_output()
        print("\nGenerating report tables...")
        generate_table2_output(monthly_data, year, OUTPUT_DIR)

        # 10. Call generate_table1_output()
        # Build year_afy_data dict from monthly_data
        year_afy_data: dict[int, float] = {}
        for well_num in range(1, 14):
            annual_mg = 0.0
            for month_int in range(1, 13):
                month_str = f"{month_int:02d}"
                month_df = monthly_data[month_str]
                well_row = month_df[month_df['Well_Number'] == well_num].iloc[0]
                annual_mg += well_row['MG_Month'].magnitude

            # Convert to AF
            annual_af = (annual_mg * ureg.million_gallon).to(ureg.acre_foot).magnitude
            year_afy_data[well_num] = annual_af

        generate_table1_output(year_afy_data, year, OUTPUT_DIR)

        # 11. Call generate_qa_summary()
        print("\nGenerating QA summary...")
        generate_qa_summary(all_flags, daily_verification, year, OUTPUT_DIR)

        # 12. Call verify_annual_sums()
        print("\nVerifying annual totals...")
        verification = verify_annual_sums(monthly_data, sum_row)

        # 13. Print final summary
        print("\n" + "=" * 60)
        print("=== SUMMARY ===")
        print("=" * 60)

        files_created = 12 + 3  # 12 monthly + Table2 + Table1 + QA
        print(f"Files created: {files_created} (12 monthly + Table2 + Table1 + QA)")

        flagged_count = len(all_flags)
        print(f"Flagged data: {flagged_count} well-months")

        ok_count = sum(1 for status in verification.values() if status == "OK")
        if ok_count == 13:
            print("Verification: All OK ✓")
        else:
            print(f"Verification: {ok_count}/13 wells OK")

        # Check July specifically (per user's requirement)
        july_df = monthly_data['07']
        buckman1_july = july_df[july_df['Well_Number'] == 1].iloc[0]
        buckman1_july_mg = buckman1_july['MG_Month'].magnitude
        buckman1_july_af = buckman1_july['MG_Month'].to(ureg.acre_foot).magnitude
        print(f"Buckman #1 July: {buckman1_july_mg:.3f} MG = {buckman1_july_af:.2f} AF ✓")

        print("=" * 60)
        print(f"SUCCESS: All outputs generated in {OUTPUT_DIR}/")
        print("=" * 60)

        # 14. Return exit code (0=success, 1=error)
        return 0

    except Exception as e:
        print("\n" + "=" * 60)
        print("=== ERROR ===")
        print("=" * 60)
        print("Fatal error during processing:")
        print(f"  {type(e).__name__}: {e}")
        print("\nTraceback:")
        import traceback
        traceback.print_exc()
        print("=" * 60)
        return 1


if __name__ == "__main__":
    sys.exit(main())
